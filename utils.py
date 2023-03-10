import pandas as pd
import numpy as np
import os, sys, glob
import humanize
import re
import xlrd

import json
import itertools
#from urllib.request import urlopen
#import requests, xmltodict
import time, datetime
import math
from pprint import pprint
import gc
from tqdm import tqdm
tqdm.pandas()
import pickle

import logging
import zipfile
import warnings
import argparse

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import units
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment

from g import logger

from manual_dictionaries import gt_cols_chunks, data_chunks, data_chunks_alter, data_chunks_alter_02
from manual_dictionaries import cols_chunks_02, main_cols, dtypes_chunks_dicts, dtypes_chunks_after_dict
from manual_dictionaries import err_msg_lst

# from tk_test import check_functions_lst

def load_check_dictionaries(path_supp_dicts):
    global df_services_MGFOMS, df_services_804n, df_RM, df_MNN, df_mi_org_gos, df_mi_national
    # if not os.path.exists(supp_dict_dir):
    #     os.path.mkdir(supp_dict_dir)

    fn = 'Коды МГФОМС.xlsx'
    fn = 'Коды МГФОМС и 804н.xlsx'
    sheet_name = 'МГФОМС'
    df_services_MGFOMS = pd.read_excel(os.path.join(path_supp_dicts, fn), sheet_name = sheet_name)
    df_services_MGFOMS.rename (columns = {'COD': 'code', 'NAME': 'name'}, inplace=True)
    df_services_MGFOMS['code'] = df_services_MGFOMS['code'].astype(str)
    # print("df_services_MGFOMS", df_services_MGFOMS.shape, df_services_MGFOMS.columns)
    logger.info(f"Загружен справочник 'Услуги по реестру  МГФОМС': {str(df_services_MGFOMS.shape)}")

    sheet_name = '804н'
    df_services_804n = pd.read_excel(os.path.join(path_supp_dicts, fn), sheet_name = sheet_name, header=1)
    df_services_804n.rename (columns = {'Код услуги': 'code', 'Наименование медицинской услуги': 'name'}, inplace=True)
    # print("df_services_804n", df_services_804n.shape, df_services_804n.columns)
    logger.info(f"Загружен справочник 'Услуги по приказу 804н': {str(df_services_804n.shape)}")

    fn = 'НВМИ_РМ.xls'
    sheet_name = 'Sheet1'
    df_RM = pd.read_excel(os.path.join(path_supp_dicts, fn), sheet_name = sheet_name)
    df_RM.rename (columns = {'Код': 'code', 'Наименование': 'name'}, inplace=True)
    df_RM['code'] = df_RM['code'].astype(str)
    # print("df_RM", df_RM.shape, df_RM.columns, df_RM.dtypes)
    logger.info(f"Загружен справочник {fn}: {str(df_RM.shape)}")
    
    # path_supp_dicts_processed = 'D:/DPP/02_tkbd/data/supp_dict/processed/'
    fn_df_mi_org_gos = 'df_mi_org_gos_release_20230129_2023_02_07_1331.pickle'
    fn_df_mi_national = 'df_mi_national_release_20230201_2023_02_06_1013.pickle'
    # df_mi_org_gos = restore_df_from_pickle(path_supp_dicts_processed, fn_df_mi_org_gos)
    # df_mi_national = restore_df_from_pickle(path_supp_dicts_processed, fn_df_mi_national)
    df_mi_org_gos = restore_df_from_pickle(path_supp_dicts, fn_df_mi_org_gos)
    df_mi_national = restore_df_from_pickle(path_supp_dicts, fn_df_mi_national)
    
    

    fn = 'МНН.xlsx'
    sheet_name = 'Sheet1'
    df_MNN = pd.read_excel(os.path.join(path_supp_dicts, fn), sheet_name = sheet_name)
    df_MNN.rename (columns = {'МНН': 'mnn_standard', 
                          'Торговое наименование лекарственного препарата': 'trade_name',
                          'Лекарственная форма, дозировка, упаковка (полная)': 'pharm_form',
                         },
               inplace=True)
    # print("df_MNN", df_MNN.shape, df_MNN.columns)
    logger.info(f"Загружен справочник {fn}: {str(df_MNN.shape)}")
    return df_services_MGFOMS, df_services_804n, df_RM, df_MNN, df_mi_org_gos, df_mi_national
# df_services_MGFOMS, df_services_804n, df_RM, df_MNN, df_mi_org_gos, df_mi_national = load_check_dictionaries(path_supp_dicts)

def np_unique_nan(lst: np.array, debug = False)->np.array: # a la version 2.4
    lst_unique = None
    if lst is None or (((type(lst)==float) or (type(lst)==np.float64)) and np.isnan(lst)):
        # if debug: print('np_unique_nan:','lst is None or (((type(lst)==float) or (type(lst)==np.float64)) and math.isnan(lst))')
        lst_unique = lst
    else:
        data_types_set = list(set([type(i) for i in lst]))
        if debug: print('np_unique_nan:', 'lst:', lst, 'data_types_set:', data_types_set)
        if ((type(lst)==list) or (type(lst)==np.ndarray)):
            if debug: print('np_unique_nan:','if ((type(lst)==list) or (type(lst)==np.ndarray)):')
            if len(data_types_set) > 1: # несколько типов данных
                if list not in data_types_set and dict not in data_types_set and tuple not in data_types_set and type(None) not in data_types_set:
                    lst_unique = np.array(list(set(lst)), dtype=object)
                else:
                    lst_unique = lst
            elif len(data_types_set) == 1:
                if debug: print("np_unique_nan: elif len(data_types_set) == 1:")
                if list in data_types_set:
                    lst_unique = np.unique(np.array(lst, dtype=object))
                elif  np.ndarray in data_types_set:
                    # print('elif  np.ndarray in data_types_set :')
                    lst_unique = np.unique(lst.astype(object))
                    # lst_unique = np_unique_nan(lst_unique)
                    lst_unique = np.asarray(lst, dtype = object)
                    # lst_unique = np.unique(lst_unique)
                elif type(None) in data_types_set:
                    # lst_unique = np.array(list(set(lst)))
                    lst_unique = np.array(list(set(list(lst))))
                elif dict in  data_types_set:
                    lst_unique = lst
                    # np.unique(lst)
                elif type(lst) == np.ndarray:
                    if debug: print("np_unique_nan: type(lst) == np.ndarray")
                    if (lst.dtype.kind == 'f') or  (lst.dtype == np.float64) or  (float in data_types_set):
                        if debug: print("np_unique_nan: (lst.dtype.kind == 'f')")
                        lst_unique = np.unique(lst.astype(float))
                        # if debug: print("np_unique_nan: lst_unique predfinal:", lst_unique)
                        # lst_unique = np.array(list(set(list(lst))))
                        # if debug: print("np_unique_nan: lst_unique predfinal v2:", lst_unique)
                        # if np.isnan(lst).all():
                        #     lst_unique = np.nan
                        #     if debug: print("np_unique_nan: lst_unique predfinal v3:", lst_unique)
                    elif (lst.dtype.kind == 'S') :
                        if debug: print("np_unique_nan: lst.dtype == string")
                        lst_unique = np.array(list(set(list(lst))))
                        if debug: print(f"np_unique_nan: lst_unique 0: {lst_unique}")
                    elif lst.dtype == object:
                        if debug: print("np_unique_nan: lst.dtype == object")
                        if (type(lst[0])==str) or (type(lst[0])==np.str_) :
                            try:
                                lst_unique = np.unique(lst)
                            except Exception as err:
                                lst_unique = np.array(list(set(list(lst))))
                        else:
                            lst_unique = np.array(list(set(list(lst))))
                        if debug: print(f"np_unique_nan: lst_unique 0: {lst_unique}")
                    else:
                        if debug: print("np_unique_nan: else 0")
                        lst_unique = np.unique(lst)
                else:
                    if debug: print('np_unique_nan:','else i...')
                    lst_unique = np.array(list(set(lst)))
                    
            elif len(data_types_set) == 0:
                lst_unique = None
            else:
                # print('else')
                lst_unique = np.array(list(set(lst)))
        else: # другой тип данных
            if debug: print('np_unique_nan:','другой тип данных')
            # lst_unique = np.unique(np.array(list(set(lst)),dtype=object))
            # lst_unique = np.unique(np.array(list(set(lst)))) # Исходим из того что все елеменыт спсика одного типа
            lst_unique = lst
    if type(lst_unique) == np.ndarray:
        if debug: print('np_unique_nan: final: ', "if type(lst_unique) == np.ndarray")
        if lst_unique.shape[0]==1: 
            if debug: print('np_unique_nan: final: ', "lst_unique.shape[0]==1")
            lst_unique = lst_unique[0]
            if debug: print(f"np_unique_nan: final after: lst_unique: {lst_unique}")
            if (type(lst_unique) == np.ndarray) and (lst_unique.shape[0]==1):  # двойная вложенность
                if debug: print('np_unique_nan: final: ', 'one more', "lst_unique.shape[0]==1")
                lst_unique = lst_unique[0]
        elif lst_unique.shape[0]==0: lst_unique = None
    if debug: print(f"np_unique_nan: return: lst_unique: {lst_unique}")
    if debug: print(f"np_unique_nan: return: type(lst_unique): {type(lst_unique)}")
    return lst_unique


def find_rec_pd(df, srch_str, print_debug=False):
    rec_num = None
    for i, row in df.iterrows():
        row_values_rejoin = [' '.join((' '.join(v.split())).split('\n')) for v in row.values if type(v)==str]
        rejoin_col = ' '.join((' '.join(srch_str.split())).split('\n'))
        if srch_str in row_values_rejoin:
            if print_debug: print('found:', srch_str)
            # print(i)
            return i
    if print_debug: print('not found:', srch_str)
    return rec_num

# new
def find_col(srch_cols_lst : list, # list of list
             cols_lst: np.ndarray,
            print_debug = False):
    if print_debug: 
        pass
        # print("srch_cols_lst", srch_cols_lst)
        # print("row.values", cols_lst)
    gt_col_nums, col_nums, col_names = [], [], [] #None, None, None
    # row_values_rejoin = [' '.join((' '.join(v.split('\n'))).split(' ')).replace('Усредненая','Усредненная') # так только по оному пробелу происходит разделение
    # print(cols_lst.dtype)
    # row_values_rejoin = [' ) '.join(' ( '.join(' '.join(' '.join(v.split('\n')]).split()).split('(')).split(')')).replace('Усредненая','Усредненная').strip().lower() 
    row_values_rejoin = [' ) '.join(' ( '.join(' '.join(' '.join([vv.strip() for vv in v.split()]).split('\n')).split('(')).split(')')).replace('Усредненая','Усредненная').strip().lower() 
                             if ((type(v)==str) or (type(v)==np.str_) or (type(v)==object))  else '' # or not ((type(v)==float) or (type(v)==int))) 
                                     for v in cols_lst ]
    row_values_rejoin = [ ' '.join([vv.strip() for vv in v.split()]) for v in row_values_rejoin]
    if print_debug: print(f"find_col: row_values_rejoin: {row_values_rejoin}")
    fl_found = False
    for i, cols_l in  enumerate(srch_cols_lst):
    # for i, srch_col in  enumerate(srch_cols_lst):
        for srch_col in cols_l:
            # srch_col_rejoin = ' '.join((' '.join(srch_col.split())).split('\n')).strip().lower()
            # srch_col_rejoin = ' ) '.join(' ( '.join(' '.join(' '.join([vv.strip() for vv in srch_col.split()]).split('\n')).split('(')).split(')')).strip().lower()
            srch_col_rejoin = ' ) '.join(' ( '.join(' '.join(' '.join(srch_col.split('\n')).split()).split('(')).split(')')).strip().lower()
            srch_col_rejoin = ' '.join([vv.strip() for vv in srch_col_rejoin.split()]) # уберем лишние пробелы между словами после апдейта по скобкам
            
            if print_debug: print(f"find_col: srch_col_rejoin: '{srch_col_rejoin}'")
            if srch_col_rejoin in row_values_rejoin:
                fl_found = True
                # gt_col_num = i
                # col_num = row_values_rejoin.index(srch_col_rejoin)
                # col_name = cols_lst[col_num]
                gt_col_nums.append(i)
                col_nums.append(row_values_rejoin.index(srch_col_rejoin))
                col_names.append(str(cols_lst[row_values_rejoin.index(srch_col_rejoin)]))
                # return gt_col_num, col_name, col_num
    if fl_found and print_debug: 
        pass
        # print("srch_cols_lst", srch_cols_lst)
        # print("row.values", cols_lst)
        # print(gt_col_nums, col_nums, col_names)
    return gt_col_nums, col_nums, col_names

# work version find_col
def find_col_w(srch_cols_lst : list, # list of list
             cols_lst: np.ndarray,
            print_debug = False):
    if print_debug: 
        pass
        # print("srch_cols_lst", srch_cols_lst)
        # print("row.values", cols_lst)
    gt_col_nums, col_nums, col_names = [], [], [] #None, None, None
    # row_values_rejoin = [' '.join((' '.join(v.split('\n'))).split(' ')).replace('Усредненая','Усредненная') # так только по оному пробелу происходит разделение
    row_values_rejoin = [' '.join((' '.join(v.split('\n'))).split()).replace('Усредненая','Усредненная').strip() 
                             if type(v)==str  else ''
                                     for v in cols_lst ]
    fl_found = False
    for i, cols_l in  enumerate(srch_cols_lst):
    # for i, srch_col in  enumerate(srch_cols_lst):
        for srch_col in cols_l:
            srch_col_rejoin = ' '.join((' '.join(srch_col.split())).split('\n')).strip()
            if srch_col_rejoin in row_values_rejoin:
                fl_found = True
                # gt_col_num = i
                # col_num = row_values_rejoin.index(srch_col_rejoin)
                # col_name = cols_lst[col_num]
                gt_col_nums.append(i)
                col_nums.append(row_values_rejoin.index(srch_col_rejoin))
                col_names.append(str(cols_lst[row_values_rejoin.index(srch_col_rejoin)]))
                # return gt_col_num, col_name, col_num
    if fl_found and print_debug: 
        pass
        # print("srch_cols_lst", srch_cols_lst)
        # print("row.values", cols_lst)
        # print(gt_col_nums, col_nums, col_names)
    return gt_col_nums, col_nums, col_names 

# def find_rec_pd_by_col_names_03(file_name, df, chunk, srch_str_lst, main_cols, print_debug = False):

# new version 09.0.2023

def find_rec_pd_by_col_names_04(file_name, df, chunk, srch_str_lst, main_cols, print_debug = False, print_debug_main = False):
    row_num = None
    fl_found = False
    fl_incorrect_found = False
    fl_all_cols_found = False
    cols_found = []
    cols_num_found = []
    cols_found_incorrect = []
    cols_num_found_incorrect = []
    result_cols = []
    gt_col_nums, col_names, col_nums = [], [], []
    # not_found_cols_nums_chunks = [[],[],[]]
    not_found_cols_nums = []
    sections = ['Услуги', 'ЛП', 'РМ']
    for i, row in df.iterrows():
        # if i <24: continue
        # if i >30: break
        fl_found = False
        # for j in range(3):
        # gt_col_nums, col_names, col_nums = find_col(srch_str_lst[main_cols[0]:main_cols[1]+1], row.values, print_debug)
        gt_col_nums, col_nums, col_names = find_col(srch_str_lst[main_cols[0]:main_cols[1]+1], row.values, print_debug)
        # print(i, gt_col_num, col_name, col_num)
        if len(gt_col_nums)>0: # # найдены основные колонки
            fl_found = True
            if print_debug_main: print("найдены основные колонки:", i, gt_col_nums, col_names, col_nums)
                # break
        else: # еще одна поптыка найтипо ключевым словам
            pass
            
        if fl_found:  # найдены оснвоыне колокни
            # теперь ищем все сотальные колокни
            gt_col_nums, col_nums, col_names  = find_col(srch_str_lst, row.values, print_debug)
            if len(gt_col_nums)>0: # is not None:
                fl_found = True
                row_num = i
                if print_debug: print(row_num, gt_col_nums, col_names, col_nums)
            # print(len(gt_col_nums), len(srch_str_lst))
            if len(gt_col_nums)< len(srch_str_lst):
                not_found_cols_nums = list(set(list(range(len(srch_str_lst)))) - set(gt_col_nums))
                not_found_cols_names = [v[0] for v in np.array(srch_str_lst, dtype=object)[not_found_cols_nums]]
                # print(f"file: {file_name}, chunk: {chunk}, строка {i}: не найдены все названия колонок, а именно:", not_found_cols_nums,
                #      not_found_cols_names)
                logger.info(f"file: {file_name}") 
                logger.info(f"Раздел: {sections[chunk]}, строка {i}: не найдены все названия колонок, а именно:"+\
                           f" {str(not_found_cols_nums)}, {str(not_found_cols_names)}")
                  # list(np.array(srch_str_lst)[:,0]).index(not_found_cols_nums))
                if print_debug_main:
                    print(f"find_rec_pd_by_col_names_04: chunk: {chunk}, gt_col_nums: {gt_col_nums}, col_nums: {col_nums}")
                
                ideal_gt_col_nums = list(range(len(srch_str_lst)))
                num_diff = len(ideal_gt_col_nums) - len(gt_col_nums)
                i_num_ins = -1
                if num_diff ==1:
                    for i_num in ideal_gt_col_nums:
                        if i_num not in gt_col_nums:
                            if (i_num == 0) or (i_num==ideal_gt_col_nums[-1]): # work cersion
                            # if (i_num >= 0) and (i_num<=ideal_gt_col_nums[-1]):
                                if print_debug_main: 
                                    print(f"find_rec_pd_by_col_names_04: 'if (i_num == 0) or (i_num==ideal_gt_col_nums[-1])', i_num: {i_num}, gt_col_nums: {gt_col_nums}")
                                if (i_num == 0):
                                    col_nums.insert(i_num, i_num)
                                elif (i_num==ideal_gt_col_nums[-1]): # добавляем № последней колонки+1
                                    col_nums.append(col_nums[-1]+1)
                                gt_col_nums.insert(i_num, i_num)
                                i_num_ins = i_num
                                col_names.insert(i_num, srch_str_lst[i_num_ins][0])
                                not_found_cols_nums.remove(i_num_ins)
                            # else:
                            # для РМ например непонятно какую брать колонку 2 или 3-ю жто не просто
                            # gt_col_nums = [0, 1, 3, 4, 5, 6]
                            # col_nums = [0, 1,  4, 5, 6, 7]
#                                 # col_nums.insert(i_num, col_nums[i_num-1]+1)
#                                 i_num_ins = gt_col_nums.index(i_num-1) + 1
#                                 # i_num_ins = col_nums.index(i_num-1) + 1
#                                 iin = col_nums.index(i_num-1) + 1
#                                 i_num = i_num_ins
# #                                 # col_nums.insert(i_num_ins, col_nums[i_num-1]+1)
#                                 col_nums.insert(iin, i_num_ins) #, i_num)
#                                 gt_col_nums.insert(i_num, i_num)
                    if len(ideal_gt_col_nums) == len(gt_col_nums) and (i_num_ins>-1):
                        logger.info(f"Вставлена колонка  {i_num_ins}, {str(srch_str_lst[i_num_ins][0])}")
                    #     # print(f"find_rec_pd_by_col_names_02:", row_num, gt_col_nums, col_names, col_nums)
                    #     return row_num, gt_col_nums, col_names, col_nums, not_found_cols_nums
                    # else:
                    return row_num, gt_col_nums, col_names, col_nums, not_found_cols_nums
                else:
                    return row_num, gt_col_nums, col_names, col_nums, not_found_cols_nums
            else:
                return row_num, gt_col_nums, col_names, col_nums, not_found_cols_nums
    if len(gt_col_nums)==0:
        logger.info(f"file: {file_name}") 
        logger.info(f"Раздел: {sections[chunk]}, не найдены основные колоноки")
    return row_num, gt_col_nums, col_names, col_nums, not_found_cols_nums# new version 16.02.2023

def find_rec_pd_by_col_names_02(file_name, df, chunk, srch_str_lst, main_cols, print_debug = False):
    row_num = None
    fl_found = False
    fl_incorrect_found = False
    fl_all_cols_found = False
    cols_found = []
    cols_num_found = []
    cols_found_incorrect = []
    cols_num_found_incorrect = []
    result_cols = []
    gt_col_nums, col_names, col_nums = [], [], []
    sections = ['Услуги', 'ЛП', 'РМ']
    for i, row in df.iterrows():
        # if i <24: continue
        # if i >30: break
        fl_found = False
        # for j in range(3):
        gt_col_nums, col_names, col_nums = find_col(srch_str_lst[main_cols[0]:main_cols[1]+1], row.values, print_debug)
        # print(i, gt_col_num, col_name, col_num)
        if len(gt_col_nums)>0: # # найдены оснвоыне колокни
            fl_found = True
            if print_debug: print(i, gt_col_nums, col_names, col_nums)
                # break
        else: # еще одна поптыка найтипо ключевым словам
            pass
            
        if fl_found:  # найдены оснвоыне колокни
            # теперь ищем все сотальные колокни
            gt_col_nums, col_nums, col_names = find_col(srch_str_lst, row.values, print_debug)
            if len(gt_col_nums)>0: # is not None:
                fl_found = True
                row_num = i
                if print_debug: print(row_num, gt_col_nums, col_names, col_nums)
            # print(len(gt_col_nums), len(srch_str_lst))
            if len(gt_col_nums)< len(srch_str_lst):
                not_found_cols_nums = list(set(list(range(len(srch_str_lst)))) - set(gt_col_nums))
                not_found_cols_names = [v[0] for v in np.array(srch_str_lst, dtype=object)[not_found_cols_nums]]
                # print(f"file: {file_name}, chunk: {chunk}, строка {i}: не найдены все названия колонок, а именно:", not_found_cols_nums,
                #      not_found_cols_names)
                logger.info(f"file: {file_name}") 
                logger.info(f"Раздел: {sections[chunk]}, строка {i}: не найдены все названия колонок, а именно:"+\
                           f" {str(not_found_cols_nums)}, {str(not_found_cols_names)}")
                  # list(np.array(srch_str_lst)[:,0]).index(not_found_cols_nums))
                
                ideal_gt_col_nums = list(range(len(srch_str_lst)))
                num_diff = len(ideal_gt_col_nums) - len(gt_col_nums)
                i_num_ins = -1
                if num_diff ==1:
                    for i_num in ideal_gt_col_nums:
                        if i_num not in gt_col_nums:
                            if (i_num == 0) or (i_num==ideal_gt_col_nums[-1]): # work cersion
                            # if (i_num >= 0) and (i_num<=ideal_gt_col_nums[-1]):
                                col_nums.insert(i_num, i_num)
                                gt_col_nums.insert(i_num, i_num)
                                i_num_ins = i_num
                                col_names.insert(i_num, srch_str_lst[i_num_ins][0])
                            # else:
                            # для РМ например непонятно какую брать колонку 2 или 3-ю жто не просто
                            # gt_col_nums = [0, 1, 3, 4, 5, 6]
                            # col_nums = [0, 1,  4, 5, 6, 7]
#                                 # col_nums.insert(i_num, col_nums[i_num-1]+1)
#                                 i_num_ins = gt_col_nums.index(i_num-1) + 1
#                                 # i_num_ins = col_nums.index(i_num-1) + 1
#                                 iin = col_nums.index(i_num-1) + 1
#                                 i_num = i_num_ins
# #                                 # col_nums.insert(i_num_ins, col_nums[i_num-1]+1)
#                                 col_nums.insert(iin, i_num_ins) #, i_num)
#                                 gt_col_nums.insert(i_num, i_num)
                    if len(ideal_gt_col_nums) == len(gt_col_nums) and (i_num_ins>-1):
                        logger.info(f"Вставлена колонка  {i_num_ins}, {str(srch_str_lst[i_num_ins][0])}")
                        # print(f"find_rec_pd_by_col_names_02:", row_num, gt_col_nums, col_names, col_nums)
                        return row_num, gt_col_nums, col_names, col_nums
            else:
                return row_num, gt_col_nums, col_names, col_nums
    return row_num, gt_col_nums, col_names, col_nums
# new version  25.01.2023    
def find_rec_pd_by_col_names_02_00(file_name, df, chunk, srch_str_lst, main_cols, print_debug = False):
    row_num = None
    fl_found = False
    fl_incorrect_found = False
    fl_all_cols_found = False
    cols_found = []
    cols_num_found = []
    cols_found_incorrect = []
    cols_num_found_incorrect = []
    result_cols = []
    
    for i, row in df.iterrows():
        # if i <24: continue
        # if i >30: break
        fl_found = False
        # for j in range(3):
        gt_col_nums, col_names, col_nums = find_col(srch_str_lst[main_cols[0]:main_cols[1]+1], row.values, print_debug)
        # print(i, gt_col_num, col_name, col_num)
        if len(gt_col_nums)>0: # # найдены оснвоыне колокни
            fl_found = True
            if print_debug: print(i, gt_col_nums, col_names, col_nums)
                # break
        else: # еще одна поптыка найтипо ключевым словам
            pass
            
        if fl_found:  # найдены оснвоыне колокни
            # теперь ищем все сотальные колокни
            gt_col_nums, col_nums, col_names = find_col(srch_str_lst, row.values, print_debug)
            if len(gt_col_nums)>0: # is not None:
                fl_found = True
                row_num = i
                if print_debug: print(row_num, gt_col_nums, col_names, col_nums)
            # print(len(gt_col_nums), len(srch_str_lst))
            if len(gt_col_nums)< len(srch_str_lst):
                not_found_cols_nums = list(set(list(range(len(srch_str_lst)))) - set(gt_col_nums))
                not_found_cols_names = [v[0] for v in np.array(srch_str_lst, dtype=object)[not_found_cols_nums]]
                # print(f"file: {file_name}, chunk: {chunk}, строка {i}: не найдены все названия колонок, а именно:", not_found_cols_nums,
                #      not_found_cols_names)
                logger.info(f"file: {file_name}, chunk: {chunk}, строка {i}: не найдены все названия колонок, а именно:"+\
                           f" {str(not_found_cols_nums)}, {str(not_found_cols_names)}")
                  # list(np.array(srch_str_lst)[:,0]).index(not_found_cols_nums))
                
                ideal_gt_col_nums = list(range(len(srch_str_lst)))
                num_diff = len(ideal_gt_col_nums) - len(gt_col_nums)
                i_num_ins = -1
                if num_diff ==1:
                    for i_num in ideal_gt_col_nums:
                        if i_num not in gt_col_nums:
                            if (i_num == 0) or (i_num==ideal_gt_col_nums[-1]):
                                col_nums.insert(i_num, i_num)
                                gt_col_nums.insert(i_num, i_num)
                                i_num_ins = i_num
                                col_names.insert(i_num, srch_str_lst[i_num_ins][0])
                            # else:
                            # для РМ например непонятно какую брать колонку 2 или 3-ю жто не просто
                            # gt_col_nums = [0, 1, 3, 4, 5, 6]
                            # col_nums = [0, 1,  4, 5, 6, 7]
                            #     col_nums.insert(i_num, col_nums[i_num-1]+1)
                            #     gt_col_nums.insert(i_num, i_num)
                    if len(ideal_gt_col_nums) == len(gt_col_nums):
                        logger.info(f"Вставлена колонка  {i_num_ins}, {str(srch_str_lst[i_num_ins][0])}")
                        # print(f"find_rec_pd_by_col_names_02:", row_num, gt_col_nums, col_names, col_nums)
                        return row_num, gt_col_nums, col_names, col_nums
            else:
                return row_num, gt_col_nums, col_names, col_nums
    return row_num, gt_col_nums, col_names, col_nums

def find_rec_pd_by_col_names_02_w(file_name, df, chunk, srch_str_lst, main_cols, print_debug = False):
    row_num = None
    fl_found = False
    fl_incorrect_found = False
    fl_all_cols_found = False
    cols_found = []
    cols_num_found = []
    cols_found_incorrect = []
    cols_num_found_incorrect = []
    result_cols = []
    
    for i, row in df.iterrows():
        # if i <24: continue
        # if i >30: break
        fl_found = False
        # for j in range(3):
        gt_col_nums, col_names, col_nums = find_col(srch_str_lst[main_cols[0]:main_cols[1]+1], row.values, print_debug)
        # print(i, gt_col_num, col_name, col_num)
        if len(gt_col_nums)>0: # is not None:
            fl_found = True
            if print_debug: print(i, gt_col_nums, col_names, col_nums)
                # break
        if fl_found:  # найдены оснвоыне колокни
            # теперь ищем все сотальные колокни
            gt_col_nums, col_nums, col_names = find_col(srch_str_lst, row.values, print_debug)
            if len(gt_col_nums)>0: # is not None:
                fl_found = True
                row_num = i
                if print_debug: print(row_num, gt_col_nums, col_names, col_nums)
            # print(len(gt_col_nums), len(srch_str_lst))
            if len(gt_col_nums)< len(srch_str_lst):
                not_found_cols_nums = list(set(list(range(len(srch_str_lst)))) - set(gt_col_nums))
                not_found_cols_names = [v[0] for v in np.array(srch_str_lst, dtype=object)[not_found_cols_nums]]
                print(f"file: {file_name}, chunk: {chunk}, строка {i}: не найдены все названия колонок, а именно:", not_found_cols_nums,
                     not_found_cols_names)
                  # list(np.array(srch_str_lst)[:,0]).index(not_found_cols_nums))
        
            else:
                return row_num, gt_col_nums, col_names, col_nums
    return row_num, gt_col_nums, col_names, col_nums

# def test_extract_chunk_positions(df_tk, j, print_debug = False, print_debug_main = False):
def test_extract_chunk_positions_02(file_name, df_tk, print_debug = False, print_debug_main = False):
    chunk_positions = [[None, None, None, None, None, None], [None, None, None, None, None, None], [None, None, None, None, None, None]]
    rec_num_0 = None
    # all_cols_found = [True, True, True] # по коли-ву обрбатываемых сейчас чанков
    all_cols_found = True
    cols_are_duplicated = False
    for j in range(6): # на всякий случай бывает что не бывает лечнбного питания лиеты после МИ 
        # проверяем наличие заголовков других блоков данных
        if print_debug: 
            print("chunk:", j)
        rec_num_0 = find_rec_pd(df_tk, data_chunks[j], print_debug=print_debug)
        if rec_num_0 is not None:
            pass
            if print_debug: print(rec_num_0)
        else:
            rec_num_0 = find_rec_pd(df_tk, data_chunks_alter[j], print_debug=print_debug)
            if rec_num_0 is not None:
                pass
                if print_debug: print(rec_num_0)
            else:
                rec_num_0 = find_rec_pd(df_tk, data_chunks_alter_02[j], print_debug=print_debug)
                if rec_num_0 is not None:
                    pass
                    if print_debug: print(rec_num_0)
        if rec_num_0 is not None:
            if (j>0) and (j<3):
                chunk_positions[j-1][1]= rec_num_0
                # try:
                #     chunk_positions[j-1][1]= rec_num_0
                # except Exception as err:
                #     print(err, "chunk_positions[j-1][1]= rec_num_0", j, chunk_positions)
                #     sys.exit(2)
            elif j>=3:
                if chunk_positions[2][1] is None:
                    chunk_positions[2][1] = rec_num_0
                
            
        if j <3:
            row_num, gt_col_nums, col_names, col_nums, not_found_cols_nums = \
                find_rec_pd_by_col_names_04(file_name, df_tk, j, cols_chunks_02[j], main_cols[j], print_debug = print_debug, print_debug_main=print_debug_main)
            # if print_debug_main:
            #     print(f"test_extract_chunk_positions: chunk: {j}, row_num: {row_num}, gt_col_nums: {gt_col_nums}, col_names: {col_names}, col_nums: {col_nums}")
            # print(f"test_extract_chunk_positions: chunk: {j} row_num, gt_col_nums, col_names, col_nums", row_num, gt_col_nums, col_names, col_nums)
            if len (gt_col_nums) < len (cols_chunks_02[j]):
                if print_debug_main:
                    print(f"test_extract_chunk_positions: chunk: {j} ->len (gt_col_nums) < len (cols_chunks_02[j])")
                    print(f"test_extract_chunk_positions: chunk: {j} gt_col_nums : {gt_col_nums},  len (cols_chunks_02[j]): {len (cols_chunks_02[j])}")
                all_cols_found = False
            if (len (col_nums) > len (cols_chunks_02[j])) or (len(gt_col_nums) > len(set(gt_col_nums))):
                # the columns are duplicated
                cols_are_duplicated = True
            if row_num is not None:
                chunk_positions[j][0] = row_num+1
                chunk_positions[j][2] = col_nums
                chunk_positions[j][3] = gt_col_nums
                chunk_positions[j][4] = not_found_cols_nums
                chunk_positions[j][5] = col_names
                if j>0 and chunk_positions[j-1][1] is None:
                    chunk_positions[j-1][1] = row_num
            if print_debug: print()
            if print_debug: 
                # print("chunk:", j)
                print(row_num, gt_col_nums, col_names, col_nums)

    
    
    return chunk_positions, all_cols_found, cols_are_duplicated# new version

def test_extract_chunk_positions(file_name, df_tk, print_debug = False, print_debug_main = False):
    chunk_positions = [[None, None, None, None], [None, None, None, None], [None, None, None, None]]
    rec_num_0 = None
    # all_cols_found = [True, True, True] # по коли-ву обрбатываемых сейчас чанков
    all_cols_found = True
    cols_are_duplicated = False
    for j in range(6): # на всякий случай бывает что не бывает лечнбного питания лиеты после МИ 
        # проверяем наличие заголовков других блоков данных
        if print_debug: 
            print("chunk:", j)
        rec_num_0 = find_rec_pd(df_tk, data_chunks[j], print_debug=print_debug)
        if rec_num_0 is not None:
            pass
            if print_debug: print(rec_num_0)
        else:
            rec_num_0 = find_rec_pd(df_tk, data_chunks_alter[j], print_debug=print_debug)
            if rec_num_0 is not None:
                pass
                if print_debug: print(rec_num_0)
            else:
                rec_num_0 = find_rec_pd(df_tk, data_chunks_alter_02[j], print_debug=print_debug)
                if rec_num_0 is not None:
                    pass
                    if print_debug: print(rec_num_0)
        if rec_num_0 is not None:
            if (j>0) and (j<3):
                chunk_positions[j-1][1]= rec_num_0
                # try:
                #     chunk_positions[j-1][1]= rec_num_0
                # except Exception as err:
                #     print(err, "chunk_positions[j-1][1]= rec_num_0", j, chunk_positions)
                #     sys.exit(2)
            elif j>=3:
                if chunk_positions[2][1] is None:
                    chunk_positions[2][1] = rec_num_0
                
            
        if j <3:
            row_num, gt_col_nums, col_names, col_nums = \
                find_rec_pd_by_col_names_02(file_name, df_tk, j, cols_chunks_02[j], main_cols[j], print_debug = print_debug)
            # print(f"test_extract_chunk_positions: chunk: {j} row_num, gt_col_nums, col_names, col_nums", row_num, gt_col_nums, col_names, col_nums)
            if len (gt_col_nums) < len (cols_chunks_02[j]):
                # print(f"test_extract_chunk_positions: chunk: {j} ->len (gt_col_nums) < len (cols_chunks_02[j])")
                all_cols_found = False
            if (len (col_nums) > len (cols_chunks_02[j])) or (len(gt_col_nums) > len(set(gt_col_nums))):
                # the columns are duplicated
                cols_are_duplicated = True
            if row_num is not None:
                chunk_positions[j][0] = row_num+1
                chunk_positions[j][2] = col_nums
                chunk_positions[j][3] = gt_col_nums
                if j>0 and chunk_positions[j-1][1] is None:
                    chunk_positions[j-1][1] = row_num
            if print_debug: print()
            if print_debug: 
                # print("chunk:", j)
                print(row_num, gt_col_nums, col_names, col_nums)

    
    
    return chunk_positions, all_cols_found, cols_are_duplicated

def test_extract_chunk_positions_00(file_name, df_tk, print_debug = False, print_debug_main = False):
    chunk_positions = [[None, None, None], [None, None, None], [None, None, None]]
    rec_num_0 = None
    # all_cols_found = [True, True, True] # по коли-ву обрбатываемых сейчас чанков
    all_cols_found = True
    for j in range(6): # на всякий случай бывает что не бывает лечнбного питания лиеты после МИ 
        # проверяем наличие заголовков других блоков данных
        if print_debug: 
            print("chunk:", j)
        rec_num_0 = find_rec_pd(df_tk, data_chunks[j], print_debug=print_debug)
        if rec_num_0 is not None:
            pass
            if print_debug: print(rec_num_0)
        else:
            rec_num_0 = find_rec_pd(df_tk, data_chunks_alter[j], print_debug=print_debug)
            if rec_num_0 is not None:
                pass
                if print_debug: print(rec_num_0)
            else:
                rec_num_0 = find_rec_pd(df_tk, data_chunks_alter_02[j], print_debug=print_debug)
                if rec_num_0 is not None:
                    pass
                    if print_debug: print(rec_num_0)
        if rec_num_0 is not None:
            if (j>0) and (j<3):
                chunk_positions[j-1][1]= rec_num_0
                # try:
                #     chunk_positions[j-1][1]= rec_num_0
                # except Exception as err:
                #     print(err, "chunk_positions[j-1][1]= rec_num_0", j, chunk_positions)
                #     sys.exit(2)
            elif j>=3:
                if chunk_positions[2][1] is None:
                    chunk_positions[2][1] = rec_num_0
                
            
        if j <3:
            row_num, gt_col_nums, col_names, col_nums = \
                find_rec_pd_by_col_names_02(file_name, df_tk, j, cols_chunks_02[j], main_cols[j], print_debug = print_debug)
            # print(f"test_extract_chunk_positions: chunk: {j} row_num, gt_col_nums, col_names, col_nums", row_num, gt_col_nums, col_names, col_nums)
            if len (gt_col_nums) < len (cols_chunks_02[j]):
                # print(f"test_extract_chunk_positions: chunk: {j} ->len (gt_col_nums) < len (cols_chunks_02[j])")
                all_cols_found = False
                pass
            if row_num is not None:
                chunk_positions[j][0] = row_num+1
                chunk_positions[j][2] = col_nums
                if j>0 and chunk_positions[j-1][1] is None:
                    chunk_positions[j-1][1] = row_num
            if print_debug: print()
            if print_debug: 
                # print("chunk:", j)
                print(row_num, gt_col_nums, col_names, col_nums)

    
    
    return chunk_positions, all_cols_found

# old version test_extract_chunk_positions
def test_extract_chunk_positions_w(file_name, df_tk, print_debug = False, print_debug_main = False):
    chunk_positions = [[None, None, None], [None, None, None], [None, None, None]]
    rec_num_0 = None
    all_cols_found = True
    for j in range(4):
        if print_debug: 
            print("chunk:", j)
        rec_num_0 = find_rec_pd(df_tk, data_chunks[j], print_debug=print_debug)
        if rec_num_0 is not None:
            pass
            if print_debug: print(rec_num_0)
        else:
            rec_num_0 = find_rec_pd(df_tk, data_chunks_alter[j], print_debug=print_debug)
            if rec_num_0 is not None:
                pass
                if print_debug: print(rec_num_0)
            else:
                rec_num_0 = find_rec_pd(df_tk, data_chunks_alter_02[j], print_debug=print_debug)
                if rec_num_0 is not None:
                    pass
                    if print_debug: print(rec_num_0)
        if rec_num_0 is not None:
            if j>0:
                chunk_positions[j-1][1]= rec_num_0
            
        if j <3:
            row_num, gt_col_nums, col_names, col_nums = \
                find_rec_pd_by_col_names_02(file_name, df_tk, j, cols_chunks_02[j], main_cols[j], print_debug = print_debug)
            if len (gt_col_nums) < len (cols_chunks_02[j]):
                all_cols_found = False
            if row_num is not None:
                chunk_positions[j][0] = row_num+1
                chunk_positions[j][2] = col_nums
                if j>0 and chunk_positions[j-1][1] is None:
                    chunk_positions[j-1][1] = row_num
            if print_debug: print()
            if print_debug: 
                # print("chunk:", j)
                print(row_num, gt_col_nums, col_names, col_nums)
    return chunk_positions, all_cols_found   

def read_chunks(path_tkbd_source, fn, sheet_name, chunks_positions, print_debug=False, print_debug_main=False):
    df_chunks = [None, None, None]
    for i, chunk_positions in enumerate(chunks_positions):
        if print_debug_main: print(chunk_positions)
        # if i ==0: continue # test
        df_chunks[i] = pd.read_excel(os.path.join(path_tkbd_source, fn), sheet_name = sheet_name, 
            skiprows = chunk_positions[0], nrows = chunk_positions[1] - chunk_positions[0], #-1,
            # index_col = chunk_positions[2],
            usecols = chunk_positions[2],
            # header=None,
            names = gt_cols_chunks[i][1:len(cols_chunks_02[i])+1],
            dtype =  dtypes_chunks_dicts[i]
        )
        df_chunk = df_chunks[i]
        if print_debug_main:
            print(f"read_chunks: chunk: {i}, df_chunk.columns: {df_chunk.columns}")
        # col_npp_name = chunk_positions[5][0]
        col_npp_name = "№ п/п"
        if not(col_npp_name is None or ((type (col_npp_name)==float) and np.isnan(col_npp_name))):
            # next_col = list(df_chunk.columns)[list(df_chunk.columns).index('№ п/п')+1]
            next_col = list(df_chunk.columns)[list(df_chunk.columns).index(col_npp_name)+1]
            if print_debug: print("next_col:", next_col)
            # Согласовано
            # Согласовано: 
            # df_chunk = df_chunk[~( ((df_chunk['№ п/п']=='nan') & \
            df_chunk = df_chunk[~( ((df_chunk[col_npp_name]=='nan') & \
                        df_chunk[next_col].str.contains(r"Согласовано", case=False ))\
                      | df_chunk[col_npp_name].str.contains(r"Согласовано", case=False)\
                      | df_chunk[next_col].str.contains(r"Согласовано", case=False))] # ,case = False
                      # df_chunk[next_col].str.contains(r"Согласовано", regex=True, flags= re.I ))\
                      # | df_chunk[col_npp_name].str.contains(r"Согласовано", regex=True, flags= re.I)\
                      # | df_chunk[next_col].str.contains(r"Согласовано", regex=True, flags= re.I))] # ,case = False
            # df_chunk = df_chunk[~( (df_chunk['№ п/п'].str.contains(data_chunks[i+1],regex=True, flags= re.I) \
            #             | df_chunk['№ п/п'].str.contains(data_chunks_alter[i+1],regex=True, flags= re.I)\
            #             | df_chunk['№ п/п'].str.contains(data_chunks_alter_02[i+1],regex=True, flags= re.I)\
            df_chunk = df_chunk[~( (df_chunk[col_npp_name].str.contains(data_chunks[i+1],regex=True, flags= re.I) \
                        | df_chunk[col_npp_name].str.contains(data_chunks_alter[i+1],regex=True, flags= re.I)\
                        | df_chunk[col_npp_name].str.contains(data_chunks_alter_02[i+1],regex=True, flags= re.I)\
                        | df_chunk[next_col].str.contains(data_chunks[i+1],regex=True, flags= re.I) \
                        | df_chunk[next_col].str.contains(data_chunks_alter[i+1],regex=True, flags= re.I) \
                        | df_chunk[next_col].str.contains(data_chunks_alter_02[i+1],regex=True, flags= re.I))) ]
            df_chunk_columns = list(df_chunk.columns)
            # next_check_cols = df_chunk_columns[df_chunk_columns.index('№ п/п')+2:-1] # 'ФИО ГВС'
            next_check_cols = df_chunk_columns[df_chunk_columns.index(col_npp_name)+2:-1] # 'ФИО ГВС'
            if print_debug_main: print(next_check_cols)
            mask_check_cols_isnan = df_chunk[next_check_cols[0]].isnull()
            for col in next_check_cols[1:]:
                mask_check_cols_isnan = mask_check_cols_isnan & df_chunk[col].isnull()
            df_chunk = df_chunk[~(mask_check_cols_isnan)]

            df_chunks[i] = df_chunk
                                            
        for k,v in dtypes_chunks_after_dict[i].items():
            if print_debug: print(f"{k}->{v}")
            # df_chunks[i][k] = df_chunks[i][k].apply( lambda x: str(x).replace(',', '.'))
            df_chunks[i][k] = df_chunks[i][k].apply( lambda x: str(x).replace(',', '.'))
            try:
                df_chunks[i][k] = df_chunks[i][k].astype(v)
            except Exception as err:
                print(f"read_chunks:-->", k, err)
        # df_chunks[i].reset_index()
        # df_chunks[i].columns = gt_cols_chunks[i][1:len(cols_chunks_02[i])+1]
        # display(df_chunks[i])
        # df_chunks[i]['Наименование ТК'] = fn + '_' + sheet_name
        # print(i, len(df_chunks[i].columns), df_chunks[i].columns)
        # print(len (gt_cols_chunks[i][1:len(cols_chunks_02[i])+1]), gt_cols_chunks[1:len(cols_chunks_02[i])])
    return df_chunks #[0], df_chunk[1], df_chunk[2]

total_sheet_names = ['Услуги', 'ЛП', 'РМ' ]
def save_to_excel(df_total, total_sheet_names, save_path, fn):
    # fn = model + '.xlsx'
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_date = fn.replace('.xlsx','')  + '_' + str_date + '.xlsx'
    
    # with pd.ExcelWriter(os.path.join(path_tkbd_processed, fn_date )) as writer:  
    with pd.ExcelWriter(os.path.join(save_path, fn_date )) as writer:  
        
        for i, df in enumerate(df_total):
            df.to_excel(writer, sheet_name = total_sheet_names[i], index=False)
    return fn_date    

def get_humanize_filesize(path, fn):
    human_file_size = None
    try:
        fn_full = os.path.join(path, fn)
    except Exception as err:
        print(err)
        return human_file_size
    if os.path.exists(fn_full):
        file_size = os.path.os.path.getsize(fn_full)
        human_file_size = humanize.naturalsize(file_size)
    return human_file_size

def save_df_to_excel(df, path_to_save, fn_main, columns = None, b=0, e=None):
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn = fn_main + '_' + str_date + '.xlsx'
    logger.info(f"'{fn}' save - start ...")
    if e is None or (e <0):
        e = df.shape[0]
    if columns is None:
        df[b:e].to_excel(path_to_save + fn, index = False)
    else:
        df[b:e].to_excel(path_to_save + fn, index = False, columns = columns)
    logger.info(f"'{fn}' saved to '{path_to_save}'")
    hfs = get_humanize_filesize(path_to_save, fn)
    logger.info("Size: " + str(hfs))
    return fn   

def restore_df_from_pickle(path_files, fn_pickle):

    if fn_pickle is None:
        logger.error('Restore pickle from ' + path_files + ' failed!')
        sys.exit(2)
    if os.path.exists(os.path.join(path_files, fn_pickle)):
        df = pd.read_pickle(os.path.join(path_files, fn_pickle))
        # logger.info('Restore ' + re.sub(path_files, '', fn_pickle_с) + ' done!')
        logger.info('Restore ' + fn_pickle + ' done!')
        logger.info('Shape: ' + str(df.shape))
    else:
        # logger.error('Restore ' + re.sub(path_files, '', fn_pickle_с) + ' from ' + path_files + ' failed!')
        logger.error('Restore ' + fn_pickle + ' from ' + path_files + ' failed!')
    return df

def unzip_file(path_source, fn_zip, work_path):
    logger.info('Unzip ' + fn_zip + ' start...')

    try:
        with zipfile.ZipFile(path_source + fn_zip, 'r') as zip_ref:
            fn_list = zip_ref.namelist()
            zip_ref.extractall(work_path)
        logger.info('Unzip ' + fn_zip + ' done!')
        return fn_list[0]
    except Exception as err:
        logger.error('Unzip error: ' + str(err))
        sys.exit(2)

def print_err_messages(rez_code_values, err_msg_lst):
    for i, rez_code_sublst in enumerate(rez_code_values):
        for j, rez_code in enumerate(rez_code_sublst):
            if not rez_code:
                # print(f"i: {i}, j: {j}")
                print(err_msg_lst[i][j])
def get_err_messages(rez_code_values, err_msg_lst):
    err_messages_lst = []
    for i, rez_code_sublst in enumerate(rez_code_values):
        err_messages_lst.append([])
        for j, rez_code in enumerate(rez_code_sublst):
            # if not rez_code:
            if rez_code==0:
                # print(f"i: {i}, j: {j}")
                err_messages_lst[i].append(err_msg_lst[i][j])
    return err_messages_lst
                
    # def check_row(chunk_num, row_values, cols_num):
    #     rez_code_row, rez_message = True, None
    #     rez_code_values = []
    #     for i, f_lst in enumerate(check_functions_lst[chunk_num]):
    #         for j, f in enumerate(f_lst):
    #             if j==0:
    #                 if type(f) == tuple:
    #                     # values_lst = [row_values[cols_num[v]] for v in f[1]]
    #                     values_lst = [row_values[v] for v in f[1]]
    #                     # print(values_lst)
    #                     rez_code_values.append([f[0](values_lst)])
    #                 # print(rez_code_values)
    #                 else: #if type(f) == 'function':
    #                     # print("row_values.shape:", row_values.shape)
    #                     # rez_code_values.append([f(row_values[cols_num[i]])])
    #                     rez_code_values.append([f(row_values[i])])
                    
    #             else: 
    #                 if type(f) == tuple:
    #                     # values_lst = [row_values[cols_num[v]] for v in f[1]]  
    #                     values_lst = [row_values[v] for v in f[1]]
    #                     # print(f[0], values_lst)
    #                     # print(rez_code_values)
    #                     rez_code_values[i].append(f[0](values_lst))
    #                 else: #if type(f) == 'function':
    #                     # rez_code_values[i].append(f(row_values[cols_num[i]]))
    #                     rez_code_values[i].append(f(row_values[i]))
                    
    #     # if False in rez_code_values: rez_code_row =False
    #     # print(rez_code_values)
    #     flat_rez_code_values = [item for sublist in rez_code_values for item in sublist]
    #     # print(flat_rez_code_values)
    #     if False in flat_rez_code_values: rez_code_row =False 
    #     if 0 in flat_rez_code_values: rez_code_row = False 
        
    #     return rez_code_row, rez_code_values, #rez_message


    # def run_check_TK(data_source_dir, data_processed_dir, fn, sheet_name,
    #         tk_profile, tk_code, tk_name, patient_model,
    #         exit_at_not_all_cols = False,
    #         print_debug = False, print_debug_main = True):
        
    #     head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel', 'Название листа в файле Excel']
    #     df_tk = pd.read_excel(os.path.join(data_source_dir, fn), sheet_name= sheet_name)
    #     j = 0
    #     # chunks_positions = test_extract_chunk_positions(df_tk, j, print_debug = print_debug, print_debug_main = print_debug_main)
    #     chunks_positions, all_cols_found = test_extract_chunk_positions(fn, df_tk, print_debug = print_debug, print_debug_main = print_debug_main)
    #     chunks_positions_flat = [item for sublist in chunks_positions for item in sublist[:2]]
    #     if print_debug_main: print("chunks_positions_flat:", chunks_positions_flat)
        

    #     if None in chunks_positions_flat or not all_cols_found: 
    #         # if print_debug_main:
    #         # print(f"{fn}, {sheet_name}: Error: didn't all chunks positions find")
    #         logger.error(f"{fn}, {sheet_name}: Error: didn't find all chunks positions or all columns")
    #         logger.info(f"chunks_positions_flat: {chunks_positions_flat}")
    #         logger.info(f"all_cols_found: {all_cols_found}")
    #         if exit_at_not_all_cols:
    #             logger.info("Process finised")
    #             sys.exit(2)
    #         else:
    #             return [None, None, None]
    #     else: 

    #         if print_debug_main: print("chunks_positions:", chunks_positions)
    #         df_chunks  = read_chunks(data_source_dir, fn, sheet_name, chunks_positions, print_debug=print_debug)
    #         for i, df_chunk in enumerate(df_chunks):
    #             if print_debug_main: print("chunk:", i)
    #             chunk_num = i
    #             cols_num = chunks_positions[i][2]
    #             err_msg_lst_flat = [item for sl in err_msg_lst[i] for item in sl]
    #             # if i ==2: #continue
    #             #     display(df_chunk.head(3))
    #             for j, row in df_chunk.iterrows():
    #                 # if chunk_num==2: print(j, "row:", row)
    #                 rez_code_row, rez_code_values = check_row(i, row.values, cols_num)
    #                 # cols_num не актуально, т.к. в chunk-е все уже попорядку
                    
    #                 # rez_code_values_np = np.array([np.array(sublst, dtype=int) for sublst in rez_code_values], dtype=list)
    #                 # rez_code_values_np = np.array([sublst for sublst in rez_code_values], dtype=list)
    #                 # rez_code_values_np = rez_code_values
    #                 flat_rez_code_values = [r for sl in rez_code_values for r in sl]
    #                 flat_rez_code_values_inv = [0 if v ==1 else 1 for v in flat_rez_code_values]
    #                 # print(flat_rez_code_values)
    #                 # rez_code_values_np = np.array(rez_code_values, dtype=list)
    #                 # rez_code_values_np = np.array(flat_rez_code_values, dtype=int)
    #                 # rez_code_values_np = flat_rez_code_values
    #                 rez_code_values_np = np.array(rez_code_values, dtype=object)
    #                 flat_rez_code_values_np = np.array(flat_rez_code_values_inv, dtype=object)
    #                 # flat_rez_code_values_np_inv = [0 if v==1 else 1 for v in flat_rez_code_values ]

    #                 err_messages = get_err_messages(rez_code_values, err_msg_lst[chunk_num])
    #                 err_messages_np = [np.array(sl, dtype=object) for sl in err_messages]
    #                 # df_chunk.loc[j, ['rez_code_row', 'rez_code_values']] = np.array([check_row(i, row.values, cols_num)], dtype = object)
    #                 # df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
    #                 # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'rez_code_values_flat' ]] = \
    #                 # dict(zip(['rez_code_row','rez_code_values', 'rez_code_values_flat'],[rez_code_row, rez_code_values_np, flat_rez_code_values_np]))
    #                 df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = \
    #                 dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
    #                 # print(err_msg_lst_flat)
    #                 # print(flat_rez_code_values)

    #                 df_chunk.loc[j, err_msg_lst_flat] = dict(zip(err_msg_lst_flat, flat_rez_code_values_inv))
    #                 # df_chunk.loc[j, ['err_messages' ]] = dict(zip(['err_messages'],err_messages_np))
    #                 # df_chunk.loc[j, 'err_messages' ] = np.array(err_messages_np, dtype=object)
    #                 # df_chunk.loc[j, 'err_messages' ] = err_messages
    #                 # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'err_messages']] = \
    #                 #         [rez_code_row, rez_code_values, err_messages]
    #                 # df_chunk.loc[j, 'rez_code_row'] = rez_code_row
    #                 # df_chunk.loc[j, 'rez_code_values'] = {'rez_code_values': rez_code_values_np}
    #                 # df_chunk.loc[j, 'err_messages'] = err_messages
    #                 # dict({'rez_code_row':rez_code_row, 'rez_code_values':rez_code_values, 'err_messages':err_messages})
    #             # df_chunk[['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента']] = tk_profile, tk_code, tk_name, patient_model
    #             df_chunk['Профиль'] = tk_profile
    #             df_chunk['Код ТК'] = tk_code
    #             df_chunk['Наименование ТК'] = tk_name
    #             df_chunk['Модель пациента'] = patient_model
    #             df_chunk['Файл Excel'] = fn
    #             df_chunk['Название листа в файле Excel'] = sheet_name
    #             df_chunk_columns = list(df_chunk.columns)
    #             for col in head_cols:
    #                 df_chunk_columns.remove(col)
    #             df_chunks[i] = df_chunk[head_cols + df_chunk_columns]

    #     # fn_save = save_to_excel(df_chunks, total_sheet_names, path_tkbd_processed, 'test_' + fn)
    #     # fn_save = save_to_excel(df_chunks, total_sheet_names, data_processed_dir, 'test_' + fn)
    #     return df_chunks

    # def run_check_by_desc(data_root_dir,fn_tk_desc, data_source_dir, data_processed_dir,
    #                     print_debug = False, print_debug_main = True):
    #     df_tk_description = pd.read_excel(os.path.join(data_root_dir, fn_tk_desc))
    #     df_tk_description.head(2)

    #     # for i, fn in enumerate(fn_lst[12:13]):
    #     df_total = [None, None, None]
    #     stat_tk = []
    #     # for i, fn in enumerate(fn_lst[:]):
    #     for i, row in tqdm(df_tk_description.iterrows(), total=df_tk_description.shape[0]):
    #         # if not os.path.isfile(os.path.join(path_tkbd_source_alter, fn)) or '.xlsx' not in fn.lower(): 
    #         #     continue
    #         if 'Файл Excel' in df_tk_description.columns:
    #             fn = row['Файл Excel']
    #         else:
    #             logger.error('В описнаии нет названий файлов')
    #             sys.exit(2)
    #         if 'Название листа в файле Excel' in df_tk_description.columns:
    #             sheet_name = row['Название листа в файле Excel']
    #         else:
    #             logger.error('В описнаии нет названий листов Excel')
    #             sys.exit(2)
    #         if 'Код' in df_tk_description.columns:
    #             tk_code = row['Код']
    #         else: tk_code = None
    #         if 'Профиль' in df_tk_description.columns:
    #             tk_profile = row['Профиль']
    #         else: tk_profile = None
    #         if 'Наименование' in df_tk_description.columns:
    #             tk_name = row['Наименование']
    #         else: tk_name = None
    #         if 'Модель пациента' in df_tk_description.columns:
    #             patient_model = row['Модель пациента']
    #         else: patient_model = None
            
            
    #         if print_debug_main: 
    #             print()
    #             print(fn, sheet_name)
    #         df_chunks = run_check_TK(data_source_dir, data_processed_dir, fn, sheet_name,
    #             tk_code, tk_profile, tk_name, patient_model,
    #             print_debug = print_debug, print_debug_main = print_debug_main)
            
    #         if i == 0: 
    #             df_total = df_chunks
    #         else:
    #             for ii, df_chunk in enumerate(df_chunks):
    #                 df_total[ii] = pd.concat([df_total[ii], df_chunk])
    #         # k += 1
    #         stat_tk.append( [tk_profile, tk_code, tk_name, fn, sheet_name, 
    #                 df_chunks[0].shape[0], df_chunks[1].shape[0], df_chunks[2].shape[0]])
    #             # print()

    #     if df_total[0] is not None: 
    #         print(df_total[0].shape)
    #         total_sheet_names = ['Услуги', 'ЛП', 'РМ' ]
    #         # fn_save = save_to_excel(df_total, total_sheet_names, path_tkbd_processed, 'tkbd.xlsx')
    #         fn_save = save_to_excel(df_total, total_sheet_names, data_processed_dir, 'tkbd_check.xlsx')
    #         # str_date = fn_save.replace('.xlsx', '').split('_')[-4:])
    #         df_stat_tk = pd.DataFrame(stat_tk, columns = ['tk_profile', 'tk_code', 'tk_name', 'fn', 'sheet_name', 'Услуги', 'ЛП', 'РМ'])
    #         fm_stat_save = save_to_excel([df_stat_tk], 
    #                     ['Shapes'], data_processed_dir, 'tkbd_check_stat.xlsx')
    #     else: 
    #         fn_save = None
    #         fm_stat_save = None
    #     logger.info(f"Check file '{fn_save}' saved in '{data_processed_dir}'")
    #     logger.info(f"Check stat file '{fm_stat_save}' saved in '{data_processed_dir}'")
    #     return fn_save, fm_stat_save

    # def run_check_by_files(data_source_dir, data_processed_dir,
    #                     print_debug = False, print_debug_main = True):
    #     df_total = [None, None, None]
    #     stat_tk = []
    #     fn_lst = os.listdir(data_source_dir)
    #     k = 0
        
    #     for i, fn in tqdm(enumerate(fn_lst[:]), total = len(fn_lst)):
        
    #         if not os.path.isfile(os.path.join(data_source_dir, fn)) or '.xlsx' not in fn.lower(): 
    #             logger.info(f"file '{fn}' not found or not xlsx-file")
    #             continue
    #         tk_profile = None
    #         tk_code = None
    #         tk_name = None #re.sub(r"^\d+\.", '', fn.split(' +')[0].replace('.xlsx','')).strip()
    #         patient_model = None
    #         xl = pd.ExcelFile(os.path.join(data_source_dir, fn))
    #         xl_sheet_names = xl.sheet_names  # see all sheet names
    #         print(fn, xl_sheet_names)
    #         for sheet_name in xl_sheet_names:

    #             df_tk = pd.read_excel(os.path.join(data_source_dir, fn), sheet_name= sheet_name)

    #             print(k, sheet_name)
        
    #             # logger.error('В описнаии нет названий листов Excel')
    #             # sys.exit(2)
            
            
    #             if print_debug_main: 
    #                 print()
    #                 print(fn, sheet_name)
    #             df_chunks = run_check_TK(data_source_dir, data_processed_dir, fn, sheet_name,
    #                 tk_code, tk_profile, tk_name, patient_model,
    #                 exit_at_not_all_cols=False,
    #                 print_debug = print_debug, print_debug_main = print_debug_main)
    #             if df_chunks[0] is None : continue
                
    #             if k == 0: 
    #                 df_total = df_chunks
    #             else:
    #                 for ii, df_chunk in enumerate(df_chunks):
    #                     df_total[ii] = pd.concat([df_total[ii], df_chunk])
    #             k += 1
    #             stat_tk.append( [tk_profile, tk_code, tk_name, fn, sheet_name, 
    #                     df_chunks[0].shape[0], df_chunks[1].shape[0], df_chunks[2].shape[0]])
    #                 # print()

    #         if df_total[0] is not None: 
    #             print(df_total[0].shape)
    #             total_sheet_names = ['Услуги', 'ЛП', 'РМ' ]
    #             # fn_save = save_to_excel(df_total, total_sheet_names, path_tkbd_processed, 'tkbd.xlsx')
    #             fn_save = save_to_excel(df_total, total_sheet_names, data_processed_dir, 'tkbd_check.xlsx')
    #             # str_date = fn_save.replace('.xlsx', '').split('_')[-4:])
    #             df_stat_tk = pd.DataFrame(stat_tk, columns = ['tk_profile', 'tk_code', 'tk_name', 'fn', 'sheet_name', 'Услуги', 'ЛП', 'РМ'])
    #             fm_stat_save = save_to_excel([df_stat_tk], 
    #                         ['Shapes'], data_processed_dir, 'tkbd_check_stat.xlsx')
    #         else: 
    #             fn_save = None
    #             fm_stat_save = None
    #     logger.info(f"Check file '{fn_save}' saved in '{data_processed_dir}'")
    #     logger.info(f"Check stat file '{fm_stat_save}' saved in '{data_processed_dir}'")
    #     return fn_save, fm_stat_save    

def conv_str_lst_2_int_lst(str_lst):
    # print(str_lst)
    str_lst_w = [l.replace('[','').replace(']','') for l in str_lst]
    # print(str_lst_w)
    str_lst_w = [l.split(',') for l in  str_lst_w]
    # print(str_lst_w)
    int_lst = [[int(i.strip()) for i in l]  for l in str_lst_w] 
    return int_lst

def transform_list_form_xlsx(s):
    # print(s.split('list('))
    lst = s.split('list')
    # print(lst)
    lst = [l.replace(']) ', '').replace('])', '').replace(']', '').replace('([', '').replace('\n', '') for l in lst]
    lst = [l for l in lst if l != '[']
    # print(lst)
    return lst

def add_check_comments(path_tkbd_processed, fn_save):
    wb = load_workbook(os.path.join(path_tkbd_processed, fn_save))
    cols_wdth_lst = [[5,20,70,10,15,15,10,10], [5,20,15,25,15,15,10,10], [5,70,15,15,15,15,10,10]]
    desc_cols_num = 6
    col_num_check_row_total_lst = [8+desc_cols_num, 8+desc_cols_num, 7+desc_cols_num]
    col_num_check_row_codes_lst = [9+desc_cols_num, 9+desc_cols_num, 8+desc_cols_num]

    print(wb.sheetnames)
    for chunk_num, ws_title in enumerate(wb.sheetnames):
        ws = wb[ws_title] #wb['Услуги']
        # chunk_num = 0
        # if chunk_num==2: continue

        col_num_check_row_total = col_num_check_row_total_lst[chunk_num] #8
        col_num_check_row_codes = col_num_check_row_codes_lst[chunk_num] #9

        alignment=Alignment(horizontal='left', #'general',
                             vertical= 'top', #'bottom',
                             text_rotation=0,
                             wrap_text=True,
                             shrink_to_fit=False,
                             indent=0)
        cols_wdth = cols_wdth_lst[chunk_num] #[5,20,70,10,15,15,10,10]
        ws.auto_filter.ref = "A1:X1"
        for ir, row in enumerate(ws.values):
            # print(ir) #, row)
            # print(type(row[col_num_check_row_total]), row[col_num_check_row_total])
            # if (row[col_num_check_row_total]=='FALSE') and (ir>0):
            if ir==0:
                for ic, _ in enumerate(row):
                    cell = ws.cell(row=ir+1, column=ic+1 + desc_cols_num)
                    cell.comment = None
                    cell.alignment = alignment
                    if ic < len(cols_wdth):
                        ws.column_dimensions[cell.column_letter].width = cols_wdth[ic]
                    else: ws.column_dimensions[cell.column_letter].width = 15

            if (not row[col_num_check_row_total]) and (ir>0):
                # for ic,value in enumerate(row):
                #     print(ic, value)
                # print(type(row[col_num_check_row_codes]), row[col_num_check_row_codes])
                s = row[col_num_check_row_codes]
                str_lst = transform_list_form_xlsx(s)
                # print(str_lst)
                rez_code_values = conv_str_lst_2_int_lst(str_lst)
                # rez_code_values = conv_str_lst_2_int_lst(transform_list_form_xlsx(row[col_num_check_row_codes]))
                # print(rez_code_values)
                err_messages = get_err_messages(rez_code_values, err_msg_lst[chunk_num])
                # print(err_messages)
                for ic, err_msg_sl in enumerate(err_messages):
                    # print('->len', len(err_msg_sl))
                    comment = None
                    # cell = None
                    cell = ws.cell(row=ir+1, column=ic+1 + desc_cols_num)
                    if len(err_msg_sl)>0:
                        comment = Comment('\n'.join(err_msg_sl), "test")
                        # print(f"ic: {ic}, ir: {ir}")
                        # print(comment)
                        comment.width = 300
                        # comment.height = 50* len(err_msg_sl)
                        comment.height = 100
                        # ws["A1"].comment = comment
                        # cell = ws.cell(row=ir+1, column=ic+1)
                        # print(f"cell.coordinate: {cell.coordinate}")
                        cell.comment = comment
                        cell.fill = PatternFill('solid', fgColor="faf080")
                        # ws.cell(row=ir+1, column=ic+1, comment= comment)
                    else:
                        cell.comment = None
                        cell.fill = PatternFill('solid', fgColor="ffffff")
            # else:



            # if ir>20: break
    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_ch_com_save = 'tkbd_check_commented_' + str_date + '.xlsx'
    wb.save(os.path.join(path_tkbd_processed, fn_ch_com_save))    
    logger.info(f" file '{fn_ch_com_save}' save in '{path_tkbd_processed}'")
    