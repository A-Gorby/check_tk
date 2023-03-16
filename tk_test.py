import pandas as pd
import numpy as np
import os, sys, glob
import humanize
import re
import xlrd

import json
import itertools
#from urllib.request import urlopen
#import requests, xmltodict
import time, datetime
import math
from pprint import pprint
import gc
from tqdm import tqdm
tqdm.pandas()
import pickle

import logging
import zipfile
# import warnings
import argparse

import warnings
warnings.filterwarnings("ignore")

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.utils import units
from openpyxl.styles import Border, Side, PatternFill, GradientFill, Alignment

from g import logger

from manual_dictionaries import gt_cols_chunks, data_chunks, data_chunks_alter, data_chunks_alter_02
from manual_dictionaries import cols_chunks_02, main_cols, dtypes_chunks_dicts, dtypes_chunks_after_dict
from manual_dictionaries import err_msg_lst

# from utils import find_rec_pd, find_col, find_rec_pd_by_col_names_02, test_extract_chunk_positions
from utils import read_chunks, save_to_excel, load_check_dictionaries
from utils import conv_str_lst_2_int_lst, transform_list_form_xlsx, add_check_comments
from utils import find_rec_pd, find_col, find_rec_pd_by_col_names_04, read_chunks, test_extract_chunk_positions_02
from utils import print_err_messages, get_err_messages #, check_row
# from utils import run_check_TK, run_check_by_desc, run_check_by_files

# from utis import check_isnull, check_point_in_value, check_code_MGFOMS


def check_isnull(val):
    if val is None or (type(val)==float) and np.isnan(val): return 0 #False
    return 1 #True
def check_point_in_value(val):
    if (type(val)==str) and '.' in val: return 0 # False
    return 1
    # return True
def check_code_MGFOMS(val):
    try:
        if df_services_MGFOMS.query(f"code == '{val}'").shape[0] >0: 
            return 1
        else: return 0
    except:
        # как вариант ValueError: multi-line expressions are only valid in the context of data, use DataFrame.eval
        return 0

def check_code_804n(val):
    # print(f"check_code_804n: {type(val)}, val: {val}") 
    try:
        if df_services_804n.query(f"code == '{val}'").shape[0] >0: 
            return 1
        else: return 0
    except:
        # try:
        #     if df_services_804n.query(f"code == '''{val}'''", engine='python').shape[0] >0: 
        #         return 1
        #     else: return 0
        # except:
        return 0
        
def check_name_804n(val):
    # print(f"check_code_804n: {type(val)}, val: {val}") 
    try:
        if df_services_804n.query(f"name == '{val}'").shape[0] >0: 
            return 1
        else: return 0
    except:
        # try:
        # print(f"check_name_804n: val: {val}")
        # if df_services_804n.query(f"name == '''{val}'''", engine='python').shape[0] >0: 
        #     return 1
        # else: return 0
        # except:
        return 0
def check_LF(val):    
    if (type(val)==str):
        if '\n' in val:
            return 0
        else: return 1
    return 1

def check_rus_char(val):
    if (type(val)==str):
        for ch in val:
            if ord(ch)>=ord('А'): return 1 #True
    return 0 #False
def check_code_rus(val):
    
    if (type(val)==str):
        for ch in val:
            if ord(ch)>=ord('А'): return 0
    return 1

def check_matching_code_804n_name_serv(val_lst):
    code, name = val_lst
    try:
        if df_services_804n.query(f"code == '{code}' & name =='{name}'" ).shape[0] >0: 
            return 1
        else: return 0
    except:
        return 0

def check_matching_code_MGFOMS_name_serv(val_lst):
    code, name = val_lst
    try:
        if df_services_MGFOMS.query(f"code == '{code}' & name =='{name}'" ).shape[0] >0: 
            return 1
        else: return 0
    except:
        return 0

def check_UET1_code_MGFOMS(val_lst):
    code_MGFOMS, uet1 = val_lst
    try:
        if df_services_MGFOMS.query(f"code == '{code_MGFOMS}' & UET1 =={uet1}" ).shape[0] >0: 
            return 1
        else: return 0
    except:
        return 0

def check_UET2_code_MGFOMS(val_lst):
    code_MGFOMS, uet2 = val_lst
    try:
        if df_services_MGFOMS.query(f"code == '{code_MGFOMS}' & UET2 =={uet2}" ).shape[0] >0: 
            return 1
        else: return 0
    except:
        return 0

def check_serv_freq_bounds(val):
    # print("check_serv_freq_bounds:", type(val))
    if (type(val)==float) or (type(val) ==int):
        if (val > 0) and (val <= 1):
            return 1
        else: return 0
    elif type(val)==str:
        try:
            # prnt("try")
            val_f = float(val)
            if (val_f > 0) and (val_f <= 1):
                return 1
            else: return 0
        except:
            # print("except")
            return 0
            
    return 0
def check_float_format (val):
    try:
        _ = float(val)
        return 1
    except:
        return 0
def check_multiple_serv(val):
    # проверка кратности Услуг
    try:
        val_f= float(val)
        if (val_f >= 1) and (val_f.is_integer()):
            return 1
        else: return 0
    except:
        return 0

def check_multiple_LP(val):
    # проверка кратности ЛП
    try:
        val_f = float(val)
        if (val_f >= 1):
            return 1
        else: return 0
    except:
        return 0
    
def check_MNN(val):
    try:
        if df_MNN.query(f"mnn_standard == '{val}'" ).shape[0] >0: 
            return 1
        else: return 0
    except:
        return 0

def check_RM_name (val):
    # df_mi_org_gos, df_mi_national
    try:
        if df_mi_org_gos.query(f"name_clean == '{val}'" ).shape[0] >0: 
            return 1
        elif df_mi_national.query(f"name == '{val}'" ).shape[0] >0: 
            return 1
        else: return 0
    except:
        return 0

def check_RM_name_old (val):
    try:
        if df_RM.query(f"name == '{val}'" ).shape[0] >0: 
            return 1
        else: return 0
    except:
        return 0
    
def check_RM_code(val):
    # df_mi_org_gos, df_mi_national
    try:
        if df_mi_org_gos.query(f"kind == '{val}'" ).shape[0] >0: 
            return 1
        elif df_mi_national.query(f"code == '{val}'" ).shape[0] >0: 
            return 1
        else: return 0
    except: return 0
    

def check_RM_code_old(val):
    try:
        if df_RM.query(f"code == '{val}'" ).shape[0] >0: 
            return 1
        else: return 0
    except: return 0

def check_multiple_RM(val):
    # проверка кратности МИ/РМ
    try:
        val_f = float(val)
        if (val_f <= 1) and (val_f > 0):
            return 1
        else: return 0
    except:
        return 0

def check_empty(val):
    if val is None or (type(val)==float) and np.isnan(val): return 0 #False
    return 1 #True

check_functions_lst =[
[[check_isnull], 
 [check_isnull, check_code_804n, check_code_rus],  
 [check_isnull, check_name_804n, check_LF, (check_matching_code_804n_name_serv,[1,2])], 
 [check_isnull, check_code_MGFOMS, ], #(check_matching_code_MGFOMS_name_serv, [2,3])], 
 [check_isnull, check_point_in_value, check_serv_freq_bounds, check_float_format ], 
 [check_isnull, check_point_in_value, check_multiple_serv, check_float_format], 
 [check_isnull, (check_UET1_code_MGFOMS, [3,6])],
 [check_isnull, (check_UET2_code_MGFOMS, [3,7])], ],
# [['Не заполнено поле "№ п/п"'],
#  ["Не заполнено поле 'МНН'", "'МНН' отсутствует в справочнике"],
#  ["Не заполнено поле 'Код АТХ'", "В поле 'АТХ' ошибочно указаны русские буквы", ],
#  ["Не заполнено поле 'Форма выпуска ЛП'"],
#  ["Не заполнено поле 'Частота'",  "В поле 'Частота' стоит '.', а не ','",
#   'Частота > 1 или <= 0 или имеет недопустимый формат числа', 
    # "В поле 'Частота' недопустимый формат числа",],
#  ["Не заполнено поле 'Кратность'",  "В поле 'Кратность' стоит '.', а не ','",
#  "Кратность ошибочно < 1 или имеет недопустимый формат числа", "В поле 'Кратность' недопустимый формат числа",],
#  ["Не заполнено поле 'Единица измерения'",], # 'Ошибка в записе единиц измерения',],
#  ["Не заполнено поле 'Количество'", "В поле 'Количество' стоит '.', а не ','",
#  "В поле 'Количество' недопустимый формат числа",],

 [[check_isnull], 
 [check_isnull, check_MNN], 
 [check_isnull, check_code_rus], 
 [check_isnull], 
 [check_isnull, check_point_in_value, check_serv_freq_bounds, check_float_format], 
 [check_isnull, check_point_in_value, check_multiple_LP, check_float_format], 
 [check_isnull], 
 [check_isnull, check_point_in_value, check_float_format],  ],

# [['Не заполнено поле "№ п/п"'],
#  ['Не заполнено поле "Наименование МИ/РМ"', "'МИ' отсутствует в справочнике"],
#  ['Не заполнено поле "Код МИ/РМ"', "'Код МИ' отсутствует в справочнике"],
#  ["Не заполнено поле 'Частота'", "В поле 'Частота' стоит '.', а не ','", 
#  'Частота > 1 или <= 0 или имеет недопустимый формат числа',
#   "В поле 'Частота' недопустимый формат числа",],
#  ["Не заполнено поле 'Кратность'",  "В поле 'Кратность' стоит '.', а не ','",
#  # "Кратность ошибочно < 1 или имеет недопустимый формат числа", 
#   "Кратность ошибочно > 1 или имеет недопустимый формат числа", 
#   "В поле 'Кратность' недопустимый формат числа",],
#   ["Не заполнено поле 'Единица измерения'",], 
#  ["Не заполнено поле 'Количество'", "В поле 'Количество' стоит '.', а не ','",
#  "В поле 'Количество' недопустимый формат числа",],
# ]    
    
[[check_isnull], 
 [check_isnull, check_RM_name], 
 [check_isnull, check_RM_code], 
 [check_isnull, check_point_in_value, check_serv_freq_bounds, check_float_format], 
 [check_isnull, check_point_in_value, check_multiple_RM, check_float_format], 
 [check_isnull], 
 [check_isnull, check_point_in_value, check_float_format],  ],
    
]



def check_row(chunk_num, row_values, cols_num):
    rez_code_row, rez_message = True, None
    rez_code_values = []
    for i, f_lst in enumerate(check_functions_lst[chunk_num]):
        for j, f in enumerate(f_lst):
            if j==0:
                if type(f) == tuple:
                    # values_lst = [row_values[cols_num[v]] for v in f[1]]
                    values_lst = [row_values[v] for v in f[1]]
                    # print(values_lst)
                    rez_code_values.append([f[0](values_lst)])
                # print(rez_code_values)
                else: #if type(f) == 'function':
                    # print("row_values.shape:", row_values.shape)
                    # rez_code_values.append([f(row_values[cols_num[i]])])
                    rez_code_values.append([f(row_values[i])])
                
            else: 
                if type(f) == tuple:
                    # values_lst = [row_values[cols_num[v]] for v in f[1]]  
                    values_lst = [row_values[v] for v in f[1]]
                    # print(f[0], values_lst)
                    # print(rez_code_values)
                    rez_code_values[i].append(f[0](values_lst))
                else: #if type(f) == 'function':
                    # rez_code_values[i].append(f(row_values[cols_num[i]]))
                    rez_code_values[i].append(f(row_values[i]))
                
    # if False in rez_code_values: rez_code_row =False
    # print(rez_code_values)
    flat_rez_code_values = [item for sublist in rez_code_values for item in sublist]
    # print(flat_rez_code_values)
    if False in flat_rez_code_values: rez_code_row =False 
    if 0 in flat_rez_code_values: rez_code_row = False 
    
    return rez_code_row, rez_code_values, #rez_message

def extract_significant_cols(fn, df_tk, chunk_positions, print_debug = False, print_debug_main = False):
    significant_col_nums = []
    col_nums = chunk_positions[2]
    start_row_num = chunk_positions[0]-1 if chunk_positions[0]>0 else 0
    row_values = df_tk.iloc[start_row_num].values
    if print_debug_main: print(f"extract_significant_cols: row_values: {row_values}")
    # significant_col_nums = [i for i,v in enumerate(row_values)]
    significant_col_nums = [i for i,v in enumerate(row_values) if not( v is  None or((type(v)==float) and np.isnan(v)))]
    significant_col_names = list(row_values)
    
    return significant_col_nums, significant_col_names

# chunks_positions
# [[39, 148, [0, 1, 2, 3, 4, 5, 6, 7], [0, 1, 2, 3, 4, 5, 6, 7], [], col_names], [150, 227, [1, 3, 4, 5, 6, 7], [1, 3, 4, 5, 6, 7], [0, 2], col_names], [229, 347, [0, 1, 2, 4, 5, 6, 7], [0, 1, 2, 3, 4, 5, 6], [], col_names]]
def try_insert_columns(fn, df_tk, 
           chunks_positions, all_cols_found, cols_are_duplicated,
           print_debug = False, print_debug_main = False):
    # На входе all_cols_found is not True 
    for i_chunk, chunk_positions in  enumerate(chunks_positions):
        gt_columns_not_found = chunk_positions[4]
        if gt_columns_not_found is not None and len (gt_columns_not_found)>0:
            significant_col_nums, significant_col_names = extract_significant_cols(fn, df_tk, chunk_positions, print_debug = False, print_debug_main = False)
            if print_debug_main: print(f"try_insert_columns: i_chunk: {i_chunk}, significant_col_nums: {significant_col_nums}")
            if len(significant_col_nums) == len(cols_chunks_02[i_chunk]): 
                # хороший случай - все заполненые навзание по по кол-ву соответсвуют кол-ву колонок в gt_cols_chunks[i_chunk]
                chunks_positions[i_chunk][2] = significant_col_nums
                chunks_positions[i_chunk][3] = list(range(len(cols_chunks_02[i_chunk]))) 
                chunks_positions[i_chunk][4] = []
                chunks_positions[i_chunk][5] = significant_col_names
                all_cols_found = True
            else: 
                col_nums_real = []
                col_names_real = []
                col_nums = chunk_positions[2]
                gt_col_nums = chunk_positions[3]
                col_names = chunk_positions[5]
                gt_col_nums_ideal = list(range(len(cols_chunks_02[i_chunk])))
                start_col_num = col_nums[0]
                start_gt_col_num = gt_col_nums[0]
                if print_debug_main:
                    print(f"try_insert_columns: len(significant_col_nums) <> len(gt_cols_chunks[i_chunk])")
                    print(f"try_insert_columns: col_nums: {col_nums}, gt_col_nums: {gt_col_nums}, gt_columns_not_found: {gt_columns_not_found}")
                if print_debug_main:
                    print(f"try_insert_columns: start_col_num: {start_col_num}, gt_col_nums_ideal: {gt_col_nums_ideal}")
                if gt_col_nums[0] == 0: # найденные нужные колонки начинаются с 0, с самой первой
                    # start_col_num = col_nums[0]
                    i_col_cnt = 0
                    for i_col, gt_col_num_ideal in enumerate(gt_col_nums_ideal):
                        if gt_col_num_ideal in gt_columns_not_found:
                            col_nums_real.append(significant_col_nums[start_col_num + gt_col_num_ideal]) 
                            col_names_real.append(significant_col_names[start_col_num + gt_col_num_ideal]) 
                        else:
                            # print(f"i_col: {i_col}")
                            col_nums_real.append(col_nums[i_col_cnt])
                            col_names_real.append(col_names[i_col_cnt])
                            i_col_cnt += 1
                        
                    chunks_positions[i_chunk][2] = col_nums_real
                    chunks_positions[i_chunk][3] = gt_col_nums_ideal
                    chunks_positions[i_chunk][4] = []
                    chunks_positions[i_chunk][5] = col_names_real
                    all_cols_found = True
                else: # начинается не с 1-й нужной  колонки
                    # col_nums: [1, 2, 4, 5, 6], gt_col_nums: [1, 2, 3, 4, 5], not_found_cols: [0, 6], gt_col_nums_ideal: [0, 1, 2, 3, 4, 5, 6]
                    # сначала идем от текущей к 1-ой колонке
                    i_col_cnt_reverse = start_gt_col_num-1
                    for i_col, gt_col_num_ideal in enumerate(gt_col_nums_ideal[start_gt_col_num-1::-1]):
                        i_col_reverse = start_gt_col_num -1 - i_col
                        if gt_col_num_ideal in gt_columns_not_found:
                            if print_debug_main: print(f"backwards: gt_col_num_ideal: {gt_col_num_ideal}, if gt_col_num_ideal in gt_columns_not_found:")
                            # col_nums_real.insert(i_col_reverse, significant_col_nums[start_col_num - i_col_reverse + gt_col_num_ideal]) 
                            col_nums_real.insert(0, significant_col_nums[start_col_num - i_col -1]) # + gt_col_num_ideal]) 
                            col_names_real.insert(0, significant_col_names[start_col_num - i_col -1]) # + gt_col_num_ideal]) 
                            
                        else:
                            if print_debug_main: print(f"backwards: gt_col_num_ideal: {gt_col_num_ideal}, else: if gt_col_num_ideal in gt_columns_not_found:")
                            col_nums_real.insert(0, col_nums[i_col_cnt_reverse])
                            col_names_real.insert(0, col_names[i_col_cnt_reverse])
                            i_col_cnt_reverse -= 1
                    # потом идем от текущей к последней колонке
                    i_col_cnt = 0
                    for i_col, gt_col_num_ideal in enumerate(gt_col_nums_ideal[start_gt_col_num::]):
                        if gt_col_num_ideal in gt_columns_not_found:
                            if print_debug_main: print(f"lookahead: gt_col_num_ideal: {gt_col_num_ideal}, if gt_col_num_ideal in gt_columns_not_found:")
                            # col_nums_real.insert(i_col_reverse, significant_col_nums[start_col_num - i_col_reverse + gt_col_num_ideal]) 
                            col_nums_real.append(significant_col_nums[start_col_num + gt_col_num_ideal - start_gt_col_num+1]) 
                            col_names_real.append(significant_col_names[start_col_num + gt_col_num_ideal - start_gt_col_num+1]) 
                            
                        else:
                            if print_debug_main: print(f"lookahead: gt_col_num_ideal: {gt_col_num_ideal}, else: if gt_col_num_ideal in gt_columns_not_found:")
                            col_nums_real.append(col_nums[i_col_cnt])
                            col_names_real.append(col_names[i_col_cnt])
                            i_col_cnt += 1
                    chunks_positions[i_chunk][2] = col_nums_real
                    chunks_positions[i_chunk][3] = gt_col_nums_ideal
                    chunks_positions[i_chunk][4] = []
                    chunks_positions[i_chunk][5] = col_names_real
                    all_cols_found = True
    for chunk_positions in chunks_positions:
        if chunk_positions[4] is not None and (len(chunk_positions[4])>0): #not_found_cols
            all_cols_found = False
        if chunk_positions[4] is None:
            all_cols_found = False
    return chunks_positions, all_cols_found, cols_are_duplicated

def run_check_TK_02(data_source_dir, data_processed_dir, fn, sheet_name,
         tk_profile, tk_code, tk_name, patient_model,
         exit_at_not_all_cols = False,
         print_debug = False, print_debug_main = True):
    
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel', 'Название листа в файле Excel']
    df_tk = pd.read_excel(os.path.join(data_source_dir, fn), sheet_name= sheet_name)
    j = 0
    # chunks_positions = test_extract_chunk_positions(df_tk, j, print_debug = print_debug, print_debug_main = print_debug_main)
    chunks_positions, all_cols_found, cols_are_duplicated = test_extract_chunk_positions_02(fn, df_tk, print_debug = print_debug, print_debug_main = print_debug_main)
    chunks_positions_flat = [item for sublist in chunks_positions for item in sublist[:2]]
    if print_debug_main: 
        print("chunks_positions after test_extract_chunk_positions:", chunks_positions)
        print(f"all_cols_found: {all_cols_found}, cols_are_duplicated: {cols_are_duplicated}")
        print("chunks_positions_flat:", chunks_positions_flat)
    
    if not all_cols_found:
        # logger.info(f"Лист '{sheet_name}':")
        logger.info(f"Попытка переопределения недостающих колонок...")
        chunks_positions, all_cols_found, cols_are_duplicated = try_insert_columns(fn, df_tk, 
                       chunks_positions, all_cols_found, cols_are_duplicated,
                       print_debug = print_debug, print_debug_main = print_debug_main)
        if print_debug_main: print(f"run_check_TK_02: after try_insert_columns: chunks_positions: {chunks_positions}, all_cols_found: {all_cols_found}, cols_are_duplicated: {cols_are_duplicated}")
        if all_cols_found:
            # logger.info(f"Лист '{sheet_name}':")
            logger.info(f"Недостающие колонки переопределены!")
        chunks_positions_flat = [item for sublist in chunks_positions for item in sublist[:2]]
        
    if None in chunks_positions_flat or not all_cols_found or cols_are_duplicated: 
        if print_debug_main:
        # print(f"{fn}, {sheet_name}: Error: didn't all chunks positions find")
            logger.error(f"Файл: {fn}")
            logger.error(f"Лист: {sheet_name}: Ошибка: Не найдены все разделы или все колокни разделов")
            logger.info(f"chunks_positions_flat: {chunks_positions_flat}")
            logger.info(f"all_cols_found: {all_cols_found}")
        if cols_are_duplicated:
            sections = ['Услуги', 'ЛП', 'РМ']
            section_cols_duplicated = []
            for j, (_, _, col_nums, gt_col_nums) in enumerate(chunks_positions):
                if (len (col_nums) > len (cols_chunks_02[j])) or (len(gt_col_nums) > len(set(gt_col_nums))):
                    section_cols_duplicated.append(sections[j])
                    
            logger.error(f"Колонки задублированы в разделах: '{', '.join(section_cols_duplicated)}'")
        if chunks_positions_flat[0] is None:
            # logger.info(f"Not found chunk: Услуги")
            logger.error(f"Не найден раздел: Услуги")
        if chunks_positions_flat[1] is None:
            # logger.info(f"Not found chunk: Услуги")
            logger.error(f"Не найдено завершение раздела: Услуги")
        if chunks_positions_flat[2] is None:
            logger.error(f"Не найден раздел: ЛП")
        if chunks_positions_flat[3] is None:
            logger.error(f"Не найден завершение раздела: ЛП")
        if chunks_positions_flat[4] is None:
            logger.error(f"Не найден раздел: МИ/РМ")
        if chunks_positions_flat[5] is None:
            logger.error(f"Не найден завершение раздела: МИ/РМ")
        
        if exit_at_not_all_cols:
            logger.info("Обработка завершена")
            sys.exit(2)
        # else:
        #     # logger.info(f"Файл '{fn}'")
        #     logger.error(f"Обработка листа '{sheet_name}' производиться не будет")
        #     return [None, None, None]
    # else: 
    if not cols_are_duplicated:
        # None in chunks_positions_flat
        # можем без некторых разделов
        if print_debug_main: print("chunks_positions:", chunks_positions)
        col_npp_name = '№ п/п'
        df_chunks  = read_chunks(data_source_dir, fn, sheet_name, chunks_positions, print_debug=print_debug, print_debug_main= print_debug_main)
        for i, df_chunk in enumerate(df_chunks):
            # if df_chunk is None: # колокнки по разделу не найдены, раздел не сохранился
            #     continue
            
            if print_debug_main: print("run_check_TK: chunk:", i, "process started")
            chunk_num = i
            cols_num = chunks_positions[i][2]
            err_msg_lst_flat = [item for sl in err_msg_lst[i] for item in sl]
            # if i ==2: #continue
            #     display(df_chunk.head(3))
            for j, row in df_chunk.iterrows():
                # if chunk_num==2: print(j, "row:", row)
                rez_code_row, rez_code_values = check_row(i, row.values, cols_num) #, debug=print_debug_main)
                # cols_num не актуально, т.к. в chunk-е все уже попорядку
                
                # rez_code_values_np = np.array([np.array(sublst, dtype=int) for sublst in rez_code_values], dtype=list)
                # rez_code_values_np = np.array([sublst for sublst in rez_code_values], dtype=list)
                # rez_code_values_np = rez_code_values
                flat_rez_code_values = [r for sl in rez_code_values for r in sl]
                flat_rez_code_values_inv = [0 if v ==1 else 1 for v in flat_rez_code_values]
                # print(flat_rez_code_values)
                # rez_code_values_np = np.array(rez_code_values, dtype=list)
                # rez_code_values_np = np.array(flat_rez_code_values, dtype=int)
                # rez_code_values_np = flat_rez_code_values
                rez_code_values_np = np.array(rez_code_values, dtype=object)
                flat_rez_code_values_np = np.array(flat_rez_code_values_inv, dtype=object)
                # flat_rez_code_values_np_inv = [0 if v==1 else 1 for v in flat_rez_code_values ]

                err_messages = get_err_messages(rez_code_values, err_msg_lst[chunk_num])
                err_messages_np = [np.array(sl, dtype=object) for sl in err_messages]
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values']] = np.array([check_row(i, row.values, cols_num)], dtype = object)
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'rez_code_values_flat' ]] = \
                # dict(zip(['rez_code_row','rez_code_values', 'rez_code_values_flat'],[rez_code_row, rez_code_values_np, flat_rez_code_values_np]))
                
                try:
                    # df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = \
                    #         rez_code_row, rez_code_values_np
                    # dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
                        # dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, flat_rez_code_values_np]))
                    df_chunk.loc[j, 'rez_code_row',] = rez_code_row
                    # df_chunk.loc[j, 'rez_code_values'] = rez_code_values_np
                    # df_chunk.loc[j, 'rez_code_values'] = {'rez_code_values':  np.asarray(rez_code_values_np)}
                    df_chunk.loc[j, 'rez_code_values'] = str(rez_code_values_np)
                    
                except Exception as err:
                    print(j, err, f"rez_code_row : {rez_code_row}, rez_code_values_np: {rez_code_values_np}" ) # , flat_rez_code_values_np: {flat_rez_code_values_np}")
                # print(err_msg_lst_flat)
                # print(flat_rez_code_values)
                try:
                    df_chunk.loc[j, err_msg_lst_flat] = dict(zip(err_msg_lst_flat, flat_rez_code_values_inv))
                except Exception as err:
                    print(j, err, f"err_msg_lst_flat: {err_msg_lst_flat}, flat_rez_code_values_inv: {flat_rez_code_values_inv}")
                # df_chunk.loc[j, ['err_messages' ]] = dict(zip(['err_messages'],err_messages_np))
                # df_chunk.loc[j, 'err_messages' ] = np.array(err_messages_np, dtype=object)
                # df_chunk.loc[j, 'err_messages' ] = err_messages
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'err_messages']] = \
                #         [rez_code_row, rez_code_values, err_messages]
                # df_chunk.loc[j, 'rez_code_row'] = rez_code_row
                # df_chunk.loc[j, 'rez_code_values'] = {'rez_code_values': rez_code_values_np}
                # df_chunk.loc[j, 'err_messages'] = err_messages
                # dict({'rez_code_row':rez_code_row, 'rez_code_values':rez_code_values, 'err_messages':err_messages})
            # df_chunk[['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента']] = tk_profile, tk_code, tk_name, patient_model
            df_chunk['Профиль'] = tk_profile
            df_chunk['Код ТК'] = tk_code
            df_chunk['Наименование ТК'] = tk_name
            df_chunk['Модель пациента'] = patient_model
            df_chunk['Файл Excel'] = fn
            df_chunk['Название листа в файле Excel'] = sheet_name
            df_chunk_columns = list(df_chunk.columns)
            for col in head_cols:
                df_chunk_columns.remove(col)
            df_chunks[i] = df_chunk[head_cols + df_chunk_columns]
        
        logger.info(f"Обработка листа '{sheet_name}' завершена")
    else:
            # logger.info(f"Файл '{fn}'")
            logger.error(f"Обработка листа '{sheet_name}' производиться не будет")
            return [None, None, None]
    # fn_save = save_to_excel(df_chunks, total_sheet_names, path_tkbd_processed, 'test_' + fn)
    # fn_save = save_to_excel(df_chunks, total_sheet_names, data_processed_dir, 'test_' + fn)
    return df_chunks
    

def run_check_TK_01a(data_source_dir, data_processed_dir, fn, sheet_name,
         tk_profile, tk_code, tk_name, patient_model,
         exit_at_not_all_cols = False,
         print_debug = False, print_debug_main = True):
    
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel', 'Название листа в файле Excel']
    df_tk = pd.read_excel(os.path.join(data_source_dir, fn), sheet_name= sheet_name)
    j = 0
    # chunks_positions = test_extract_chunk_positions(df_tk, j, print_debug = print_debug, print_debug_main = print_debug_main)
    chunks_positions, all_cols_found, cols_are_duplicated = test_extract_chunk_positions(fn, df_tk, print_debug = print_debug, print_debug_main = print_debug_main)
    chunks_positions_flat = [item for sublist in chunks_positions for item in sublist[:2]]
    if print_debug_main: 
        print("chunks_positions:", chunks_positions)
        print("all_cols_found, cols_are_duplicated", all_cols_found, cols_are_duplicated)
        print("chunks_positions_flat:", chunks_positions_flat)
    

    if None in chunks_positions_flat or not all_cols_found or cols_are_duplicated: 
        if print_debug_main:
        # print(f"{fn}, {sheet_name}: Error: didn't all chunks positions find")
            logger.error(f"File: {fn}")
            logger.error(f"Sheet: {sheet_name}: Error: didn't find all chunks positions or all columns")
            logger.info(f"chunks_positions_flat: {chunks_positions_flat}")
            logger.info(f"all_cols_found: {all_cols_found}")
        if cols_are_duplicated:
            sections = ['Услуги', 'ЛП', 'РМ']
            section_cols_duplicated = []
            for j, (_, _, col_nums, gt_col_nums) in enumerate(chunks_positions):
                if (len (col_nums) > len (cols_chunks_02[j])) or (len(gt_col_nums) > len(set(gt_col_nums))):
                    section_cols_duplicated.append(sections[j])
                    
            logger.info(f"Колонки задублированы в разделах: '{', '.join(section_cols_duplicated)}'")
        if chunks_positions_flat[0] is None:
            # logger.info(f"Not found chunk: Услуги")
            logger.info(f"Не найден раздел: Услуги")
        if chunks_positions_flat[1] is None:
            # logger.info(f"Not found chunk: Услуги")
            logger.info(f"Не найдено завершение раздела: Услуги")
        if chunks_positions_flat[2] is None:
            logger.info(f"Не найден раздел: ЛП")
        if chunks_positions_flat[3] is None:
            logger.info(f"Не найден завершение раздела: ЛП")
        if chunks_positions_flat[4] is None:
            logger.info(f"Не найден раздел: МИ/РМ")
        if chunks_positions_flat[5] is None:
            logger.info(f"Не найден завершение раздела: МИ/РМ")
        
        if exit_at_not_all_cols:
            logger.info("Обработка завершена")
            sys.exit(2)
        else:
            # logger.info(f"Файл '{fn}'")
            logger.info(f"Обработка листа '{sheet_name}' производиться не будет")
            return [None, None, None]
    else: 

        if print_debug_main: print("chunks_positions:", chunks_positions)
        df_chunks  = read_chunks(data_source_dir, fn, sheet_name, chunks_positions, print_debug=print_debug)
        for i, df_chunk in enumerate(df_chunks):
            if print_debug_main: print("run_check_TK: chunk:", i, "process started")
            chunk_num = i
            cols_num = chunks_positions[i][2]
            err_msg_lst_flat = [item for sl in err_msg_lst[i] for item in sl]
            # if i ==2: #continue
            #     display(df_chunk.head(3))
            for j, row in df_chunk.iterrows():
                # if chunk_num==2: print(j, "row:", row)
                rez_code_row, rez_code_values = check_row(i, row.values, cols_num) #, debug=print_debug_main)
                # cols_num не актуально, т.к. в chunk-е все уже попорядку
                
                # rez_code_values_np = np.array([np.array(sublst, dtype=int) for sublst in rez_code_values], dtype=list)
                # rez_code_values_np = np.array([sublst for sublst in rez_code_values], dtype=list)
                # rez_code_values_np = rez_code_values
                flat_rez_code_values = [r for sl in rez_code_values for r in sl]
                flat_rez_code_values_inv = [0 if v ==1 else 1 for v in flat_rez_code_values]
                # print(flat_rez_code_values)
                # rez_code_values_np = np.array(rez_code_values, dtype=list)
                # rez_code_values_np = np.array(flat_rez_code_values, dtype=int)
                # rez_code_values_np = flat_rez_code_values
                rez_code_values_np = np.array(rez_code_values, dtype=object)
                flat_rez_code_values_np = np.array(flat_rez_code_values_inv, dtype=object)
                # flat_rez_code_values_np_inv = [0 if v==1 else 1 for v in flat_rez_code_values ]

                err_messages = get_err_messages(rez_code_values, err_msg_lst[chunk_num])
                err_messages_np = [np.array(sl, dtype=object) for sl in err_messages]
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values']] = np.array([check_row(i, row.values, cols_num)], dtype = object)
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'rez_code_values_flat' ]] = \
                # dict(zip(['rez_code_row','rez_code_values', 'rez_code_values_flat'],[rez_code_row, rez_code_values_np, flat_rez_code_values_np]))
                
                try:
                    # df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = \
                    #         rez_code_row, rez_code_values_np
                    # dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
                        # dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, flat_rez_code_values_np]))
                    df_chunk.loc[j, 'rez_code_row',] = rez_code_row
                    # df_chunk.loc[j, 'rez_code_values'] = rez_code_values_np
                    # df_chunk.loc[j, 'rez_code_values'] = {'rez_code_values':  np.asarray(rez_code_values_np)}
                    df_chunk.loc[j, 'rez_code_values'] = str(rez_code_values_np)
                    
                except Exception as err:
                    print(j, err, f"rez_code_row : {rez_code_row}, rez_code_values_np: {rez_code_values_np}" ) # , flat_rez_code_values_np: {flat_rez_code_values_np}")
                # print(err_msg_lst_flat)
                # print(flat_rez_code_values)
                try:
                    df_chunk.loc[j, err_msg_lst_flat] = dict(zip(err_msg_lst_flat, flat_rez_code_values_inv))
                except Exception as err:
                    print(j, err, f"err_msg_lst_flat: {err_msg_lst_flat}, flat_rez_code_values_inv: {flat_rez_code_values_inv}")
                # df_chunk.loc[j, ['err_messages' ]] = dict(zip(['err_messages'],err_messages_np))
                # df_chunk.loc[j, 'err_messages' ] = np.array(err_messages_np, dtype=object)
                # df_chunk.loc[j, 'err_messages' ] = err_messages
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'err_messages']] = \
                #         [rez_code_row, rez_code_values, err_messages]
                # df_chunk.loc[j, 'rez_code_row'] = rez_code_row
                # df_chunk.loc[j, 'rez_code_values'] = {'rez_code_values': rez_code_values_np}
                # df_chunk.loc[j, 'err_messages'] = err_messages
                # dict({'rez_code_row':rez_code_row, 'rez_code_values':rez_code_values, 'err_messages':err_messages})
            # df_chunk[['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента']] = tk_profile, tk_code, tk_name, patient_model
            df_chunk['Профиль'] = tk_profile
            df_chunk['Код ТК'] = tk_code
            df_chunk['Наименование ТК'] = tk_name
            df_chunk['Модель пациента'] = patient_model
            df_chunk['Файл Excel'] = fn
            df_chunk['Название листа в файле Excel'] = sheet_name
            df_chunk_columns = list(df_chunk.columns)
            for col in head_cols:
                df_chunk_columns.remove(col)
            df_chunks[i] = df_chunk[head_cols + df_chunk_columns]
        
        logger.info(f"Обработка листа '{sheet_name}' завершена")

    # fn_save = save_to_excel(df_chunks, total_sheet_names, path_tkbd_processed, 'test_' + fn)
    # fn_save = save_to_excel(df_chunks, total_sheet_names, data_processed_dir, 'test_' + fn)
    return df_chunks

def run_check_TK_01(data_source_dir, data_processed_dir, fn, sheet_name,
         tk_profile, tk_code, tk_name, patient_model,
         exit_at_not_all_cols = False,
         print_debug = False, print_debug_main = True):
    
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel', 'Название листа в файле Excel']
    df_tk = pd.read_excel(os.path.join(data_source_dir, fn), sheet_name= sheet_name)
    j = 0
    # chunks_positions = test_extract_chunk_positions(df_tk, j, print_debug = print_debug, print_debug_main = print_debug_main)
    chunks_positions, all_cols_found, cols_are_duplicated = test_extract_chunk_positions(fn, df_tk, print_debug = print_debug, print_debug_main = print_debug_main)
    chunks_positions_flat = [item for sublist in chunks_positions for item in sublist[:2]]
    if print_debug_main: 
        print("chunks_positions:", chunks_positions)
        print("all_cols_found, cols_are_duplicated", all_cols_found, cols_are_duplicated)
        print("chunks_positions_flat:", chunks_positions_flat)
    

    if None in chunks_positions_flat or not all_cols_found or cols_are_duplicated: 
        if print_debug_main:
        # print(f"{fn}, {sheet_name}: Error: didn't all chunks positions find")
            logger.error(f"File: {fn}")
            logger.error(f"Sheet: {sheet_name}: Error: didn't find all chunks positions or all columns")
            logger.info(f"chunks_positions_flat: {chunks_positions_flat}")
            logger.info(f"all_cols_found: {all_cols_found}")
        if cols_are_duplicated:
            sections = ['Услуги', 'ЛП', 'РМ']
            section_cols_duplicated = []
            for j, (_, _, col_nums, gt_col_nums) in enumerate(chunks_positions):
                if (len (col_nums) > len (cols_chunks_02[j])) or (len(gt_col_nums) > len(set(gt_col_nums))):
                    section_cols_duplicated.append(sections[j])
                    
            logger.info(f"Колонки задублированы в разделах: '{', '.join(section_cols_duplicated)}'")
        if chunks_positions_flat[0] is None:
            # logger.info(f"Not found chunk: Услуги")
            logger.info(f"Не найден раздел: Услуги")
        if chunks_positions_flat[1] is None:
            # logger.info(f"Not found chunk: Услуги")
            logger.info(f"Не найдено завершение раздела: Услуги")
        if chunks_positions_flat[2] is None:
            logger.info(f"Не найден раздел: ЛП")
        if chunks_positions_flat[3] is None:
            logger.info(f"Не найден завершение раздела: ЛП")
        if chunks_positions_flat[4] is None:
            logger.info(f"Не найден раздел: МИ/РМ")
        if chunks_positions_flat[5] is None:
            logger.info(f"Не найден завершение раздела: МИ/РМ")
        
        if exit_at_not_all_cols:
            logger.info("Обработка завершена")
            sys.exit(2)
        else:
            # logger.info(f"Файл '{fn}'")
            logger.info(f"Обработка листа '{sheet_name}' производиться не будет")
            return [None, None, None]
    else: 

        if print_debug_main: print("chunks_positions:", chunks_positions)
        df_chunks  = read_chunks(data_source_dir, fn, sheet_name, chunks_positions, print_debug=print_debug)
        for i, df_chunk in enumerate(df_chunks):
            if print_debug_main: print("chunk:", i)
            chunk_num = i
            cols_num = chunks_positions[i][2]
            err_msg_lst_flat = [item for sl in err_msg_lst[i] for item in sl]
            # if i ==2: #continue
            #     display(df_chunk.head(3))
            for j, row in df_chunk.iterrows():
                # if chunk_num==2: print(j, "row:", row)
                rez_code_row, rez_code_values = check_row(i, row.values, cols_num)
                # cols_num не актуально, т.к. в chunk-е все уже попорядку
                
                # rez_code_values_np = np.array([np.array(sublst, dtype=int) for sublst in rez_code_values], dtype=list)
                # rez_code_values_np = np.array([sublst for sublst in rez_code_values], dtype=list)
                # rez_code_values_np = rez_code_values
                flat_rez_code_values = [r for sl in rez_code_values for r in sl]
                flat_rez_code_values_inv = [0 if v ==1 else 1 for v in flat_rez_code_values]
                # print(flat_rez_code_values)
                # rez_code_values_np = np.array(rez_code_values, dtype=list)
                # rez_code_values_np = np.array(flat_rez_code_values, dtype=int)
                # rez_code_values_np = flat_rez_code_values
                rez_code_values_np = np.array(rez_code_values, dtype=object)
                flat_rez_code_values_np = np.array(flat_rez_code_values_inv, dtype=object)
                # flat_rez_code_values_np_inv = [0 if v==1 else 1 for v in flat_rez_code_values ]

                err_messages = get_err_messages(rez_code_values, err_msg_lst[chunk_num])
                err_messages_np = [np.array(sl, dtype=object) for sl in err_messages]
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values']] = np.array([check_row(i, row.values, cols_num)], dtype = object)
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'rez_code_values_flat' ]] = \
                # dict(zip(['rez_code_row','rez_code_values', 'rez_code_values_flat'],[rez_code_row, rez_code_values_np, flat_rez_code_values_np]))
                df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = \
                dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
                # print(err_msg_lst_flat)
                # print(flat_rez_code_values)

                df_chunk.loc[j, err_msg_lst_flat] = dict(zip(err_msg_lst_flat, flat_rez_code_values_inv))
                # df_chunk.loc[j, ['err_messages' ]] = dict(zip(['err_messages'],err_messages_np))
                # df_chunk.loc[j, 'err_messages' ] = np.array(err_messages_np, dtype=object)
                # df_chunk.loc[j, 'err_messages' ] = err_messages
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'err_messages']] = \
                #         [rez_code_row, rez_code_values, err_messages]
                # df_chunk.loc[j, 'rez_code_row'] = rez_code_row
                # df_chunk.loc[j, 'rez_code_values'] = {'rez_code_values': rez_code_values_np}
                # df_chunk.loc[j, 'err_messages'] = err_messages
                # dict({'rez_code_row':rez_code_row, 'rez_code_values':rez_code_values, 'err_messages':err_messages})
            # df_chunk[['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента']] = tk_profile, tk_code, tk_name, patient_model
            df_chunk['Профиль'] = tk_profile
            df_chunk['Код ТК'] = tk_code
            df_chunk['Наименование ТК'] = tk_name
            df_chunk['Модель пациента'] = patient_model
            df_chunk['Файл Excel'] = fn
            df_chunk['Название листа в файле Excel'] = sheet_name
            df_chunk_columns = list(df_chunk.columns)
            for col in head_cols:
                df_chunk_columns.remove(col)
            df_chunks[i] = df_chunk[head_cols + df_chunk_columns]
        
        logger.info(f"Обработка листа '{sheet_name}' завершена")

    # fn_save = save_to_excel(df_chunks, total_sheet_names, path_tkbd_processed, 'test_' + fn)
    # fn_save = save_to_excel(df_chunks, total_sheet_names, data_processed_dir, 'test_' + fn)
    return df_chunks

def run_check_TK_00(data_source_dir, data_processed_dir, fn, sheet_name,
         tk_profile, tk_code, tk_name, patient_model,
         exit_at_not_all_cols = False,
         print_debug = False, print_debug_main = True):
    
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel', 'Название листа в файле Excel']
    df_tk = pd.read_excel(os.path.join(data_source_dir, fn), sheet_name= sheet_name)
    j = 0
    # chunks_positions = test_extract_chunk_positions(df_tk, j, print_debug = print_debug, print_debug_main = print_debug_main)
    chunks_positions, all_cols_found = test_extract_chunk_positions(fn, df_tk, print_debug = print_debug, print_debug_main = print_debug_main)
    chunks_positions_flat = [item for sublist in chunks_positions for item in sublist[:2]]
    if print_debug_main: 
        print("chunks_positions:", chunks_positions)
        print("chunks_positions_flat:", chunks_positions_flat)
    

    if None in chunks_positions_flat or not all_cols_found: 
        if print_debug_main:
        # print(f"{fn}, {sheet_name}: Error: didn't all chunks positions find")
            logger.error(f"File: {fn}")
            logger.error(f"Sheet: {sheet_name}: Error: didn't find all chunks positions or all columns")
            logger.info(f"chunks_positions_flat: {chunks_positions_flat}")
            logger.info(f"all_cols_found: {all_cols_found}")
        if chunks_positions_flat[0] is None:
            # logger.info(f"Not found chunk: Услуги")
            logger.info(f"Не найден раздел: Услуги")
        if chunks_positions_flat[2] is None:
            logger.info(f"Не найден раздел: ЛП")
        if chunks_positions_flat[4] is None:
            logger.info(f"Не найден раздел: МИ/РМ")
        
        if exit_at_not_all_cols:
            logger.info("Обработка завершена")
            sys.exit(2)
        else:
            # logger.info(f"Файл '{fn}'")
            logger.info(f"Обработка листа '{sheet_name}' производиться не будет")
            return [None, None, None]
    else: 

        if print_debug_main: print("chunks_positions:", chunks_positions)
        df_chunks  = read_chunks(data_source_dir, fn, sheet_name, chunks_positions, print_debug=print_debug)
        for i, df_chunk in enumerate(df_chunks):
            if print_debug_main: print("chunk:", i)
            chunk_num = i
            cols_num = chunks_positions[i][2]
            err_msg_lst_flat = [item for sl in err_msg_lst[i] for item in sl]
            # if i ==2: #continue
            #     display(df_chunk.head(3))
            for j, row in df_chunk.iterrows():
                # if chunk_num==2: print(j, "row:", row)
                rez_code_row, rez_code_values = check_row(i, row.values, cols_num)
                # cols_num не актуально, т.к. в chunk-е все уже попорядку
                
                # rez_code_values_np = np.array([np.array(sublst, dtype=int) for sublst in rez_code_values], dtype=list)
                # rez_code_values_np = np.array([sublst for sublst in rez_code_values], dtype=list)
                # rez_code_values_np = rez_code_values
                flat_rez_code_values = [r for sl in rez_code_values for r in sl]
                flat_rez_code_values_inv = [0 if v ==1 else 1 for v in flat_rez_code_values]
                # print(flat_rez_code_values)
                # rez_code_values_np = np.array(rez_code_values, dtype=list)
                # rez_code_values_np = np.array(flat_rez_code_values, dtype=int)
                # rez_code_values_np = flat_rez_code_values
                rez_code_values_np = np.array(rez_code_values, dtype=object)
                flat_rez_code_values_np = np.array(flat_rez_code_values_inv, dtype=object)
                # flat_rez_code_values_np_inv = [0 if v==1 else 1 for v in flat_rez_code_values ]

                err_messages = get_err_messages(rez_code_values, err_msg_lst[chunk_num])
                err_messages_np = [np.array(sl, dtype=object) for sl in err_messages]
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values']] = np.array([check_row(i, row.values, cols_num)], dtype = object)
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'rez_code_values_flat' ]] = \
                # dict(zip(['rez_code_row','rez_code_values', 'rez_code_values_flat'],[rez_code_row, rez_code_values_np, flat_rez_code_values_np]))
                df_chunk.loc[j, ['rez_code_row', 'rez_code_values' ]] = \
                dict(zip(['rez_code_row','rez_code_values'],[rez_code_row, rez_code_values_np]))
                # print(err_msg_lst_flat)
                # print(flat_rez_code_values)

                df_chunk.loc[j, err_msg_lst_flat] = dict(zip(err_msg_lst_flat, flat_rez_code_values_inv))
                # df_chunk.loc[j, ['err_messages' ]] = dict(zip(['err_messages'],err_messages_np))
                # df_chunk.loc[j, 'err_messages' ] = np.array(err_messages_np, dtype=object)
                # df_chunk.loc[j, 'err_messages' ] = err_messages
                # df_chunk.loc[j, ['rez_code_row', 'rez_code_values', 'err_messages']] = \
                #         [rez_code_row, rez_code_values, err_messages]
                # df_chunk.loc[j, 'rez_code_row'] = rez_code_row
                # df_chunk.loc[j, 'rez_code_values'] = {'rez_code_values': rez_code_values_np}
                # df_chunk.loc[j, 'err_messages'] = err_messages
                # dict({'rez_code_row':rez_code_row, 'rez_code_values':rez_code_values, 'err_messages':err_messages})
            # df_chunk[['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента']] = tk_profile, tk_code, tk_name, patient_model
            df_chunk['Профиль'] = tk_profile
            df_chunk['Код ТК'] = tk_code
            df_chunk['Наименование ТК'] = tk_name
            df_chunk['Модель пациента'] = patient_model
            df_chunk['Файл Excel'] = fn
            df_chunk['Название листа в файле Excel'] = sheet_name
            df_chunk_columns = list(df_chunk.columns)
            for col in head_cols:
                df_chunk_columns.remove(col)
            df_chunks[i] = df_chunk[head_cols + df_chunk_columns]
        
        logger.info(f"Обработка листа '{sheet_name}' завершена")

    # fn_save = save_to_excel(df_chunks, total_sheet_names, path_tkbd_processed, 'test_' + fn)
    # fn_save = save_to_excel(df_chunks, total_sheet_names, data_processed_dir, 'test_' + fn)
    return df_chunks
    
def run_check_by_desc(data_root_dir, fn_tk_desc, data_source_dir, data_processed_dir,
                     print_debug = False, print_debug_main = True):
    df_tk_description = pd.read_excel(os.path.join(data_root_dir, fn_tk_desc))
    # df_tk_description.head(2)
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel', 'Название листа в файле Excel']
    
    # for i, fn in enumerate(fn_lst[12:13]):
    df_total = [None, None, None]
    stat_tk = []
    fn_processed = []
    # for i, fn in enumerate(fn_lst[:]):
    for i, row in tqdm(df_tk_description.iterrows(), total=df_tk_description.shape[0]):
        # if not os.path.isfile(os.path.join(path_tkbd_source_alter, fn)) or '.xlsx' not in fn.lower(): 
        #     continue
        if 'Файл Excel' in df_tk_description.columns:
            fn = row['Файл Excel']
        else:
            logger.error('В описнаии нет названий файлов')
            sys.exit(2)
        if 'Название листа в файле Excel' in df_tk_description.columns:
            sheet_name = row['Название листа в файле Excel']
        else:
            logger.error('В описнаии нет названий листов Excel')
            sys.exit(2)
        if 'Код' in df_tk_description.columns:
            tk_code = row['Код']
        else: tk_code = None
        if 'Профиль' in df_tk_description.columns:
            tk_profile = row['Профиль']
        else: tk_profile = None
        if 'Наименование' in df_tk_description.columns:
            tk_name = row['Наименование']
        else: tk_name = None
        if 'Модель пациента' in df_tk_description.columns:
            patient_model = row['Модель пациента']
        else: patient_model = None
        fn_processed.append([fn, False])
        
        # if print_debug_main: 
        #     print()
        #     print(fn, sheet_name)
        logger.info(f"Файл: '{fn}'")
        logger.info(f"Лист: {str(sheet_name)}")
        df_chunks = [None, None, None]
        try: 
            df_chunks = run_check_TK_02(data_source_dir, data_processed_dir, fn, sheet_name,
                tk_code, tk_profile, tk_name, patient_model,
                 print_debug = print_debug, print_debug_main = print_debug_main)
        except Exception as err:
            logger.error(f"Файл '{fn}'")
            logger.error(f"Лист '{sheet_name}'")
            logger.error(f"Не обработан из-за ошибки: '{err}'")
            
        if i == 0: 
            df_total = df_chunks
        else:
            for ii, df_chunk in enumerate(df_chunks):
                df_total[ii] = pd.concat([df_total[ii], df_chunk])
        # k += 1
        
        # if df_chunks[0] is not None:
        #     stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
        #          df_chunks[0].shape[0], df_chunks[1].shape[0], df_chunks[2].shape[0]])
        #     fn_processed[-1][1] = True # fn_processed.append([fn, False])
        # else:
        #     stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
        #          0, 0, 0])
        row_stat = []
        for df_chunk in df_chunks:
            if df_chunk is not None:
                row_stat.append(df_chunk.shape[0])
                fn_processed[-1][1] = True # fn_processed.append([fn, False])
            else:
                row_stat.append(0)
        stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
                     row_stat[0], row_stat[1], row_stat[2]])


    if df_total[0] is not None: 
        # print(df_total[0].shape)
        total_sheet_names = ['Услуги', 'ЛП', 'РМ' ]
        # fn_save = save_to_excel(df_total, total_sheet_names, path_tkbd_processed, 'tkbd.xlsx')
        fn_save = save_to_excel(df_total, total_sheet_names, data_processed_dir, 'tkbd_check.xlsx')
        # str_date = fn_save.replace('.xlsx', '').split('_')[-4:])
        # df_stat_tk = pd.DataFrame(stat_tk, columns = ['tk_profile', 'tk_code', 'tk_name', 'fn', 'sheet_name', 'Услуги', 'ЛП', 'РМ'])
        df_stat_tk = pd.DataFrame(stat_tk, columns = head_cols + ['Услуги', 'ЛП', 'РМ'])
        # fm_stat_save = save_to_excel([df_stat_tk], 
        #               ['Shapes'], data_processed_dir, 'tkbd_check_stat.xlsx')
        df_stat_tk_files = pd.DataFrame(fn_processed, columns = ['Файл', 'Обработан'])

        fm_stat_save = save_to_excel([df_stat_tk, df_stat_tk_files], 
                      ['Shapes', 'Files'], data_processed_dir, 'tkbd_check_stat.xlsx')
    else: 
        fn_save = None
        fm_stat_save = None
    # logger.info(f"Check file '{fn_save}' saved in '{data_processed_dir}'")
    # logger.info(f"Check stat file '{fm_stat_save}' saved in '{data_processed_dir}'")
    logger.info(f"Файл проверки '{fn_save}' сохранен в '{data_processed_dir}'")
    logger.info(f"Файл статистики обработки '{fm_stat_save}' сохранен в '{data_processed_dir}'")
    
    return fn_save, fm_stat_save


def run_check_by_desc_01(data_root_dir, fn_tk_desc, data_source_dir, data_processed_dir,
                     print_debug = False, print_debug_main = True):
    df_tk_description = pd.read_excel(os.path.join(data_root_dir, fn_tk_desc))
    # df_tk_description.head(2)
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel', 'Название листа в файле Excel']
    
    # for i, fn in enumerate(fn_lst[12:13]):
    df_total = [None, None, None]
    stat_tk = []
    # for i, fn in enumerate(fn_lst[:]):
    for i, row in tqdm(df_tk_description.iterrows(), total=df_tk_description.shape[0]):
        # if not os.path.isfile(os.path.join(path_tkbd_source_alter, fn)) or '.xlsx' not in fn.lower(): 
        #     continue
        if 'Файл Excel' in df_tk_description.columns:
            fn = row['Файл Excel']
        else:
            logger.error('В описнаии нет названий файлов')
            sys.exit(2)
        if 'Название листа в файле Excel' in df_tk_description.columns:
            sheet_name = row['Название листа в файле Excel']
        else:
            logger.error('В описнаии нет названий листов Excel')
            sys.exit(2)
        if 'Код' in df_tk_description.columns:
            tk_code = row['Код']
        else: tk_code = None
        if 'Профиль' in df_tk_description.columns:
            tk_profile = row['Профиль']
        else: tk_profile = None
        if 'Наименование' in df_tk_description.columns:
            tk_name = row['Наименование']
        else: tk_name = None
        if 'Модель пациента' in df_tk_description.columns:
            patient_model = row['Модель пациента']
        else: patient_model = None
        
        
        # if print_debug_main: 
        #     print()
        #     print(fn, sheet_name)
        logger.info(f"Файл: '{fn}'")
        logger.info(f"Лист: {str(sheet_name)}")
        df_chunks = run_check_TK(data_source_dir, data_processed_dir, fn, sheet_name,
            tk_code, tk_profile, tk_name, patient_model,
             print_debug = print_debug, print_debug_main = print_debug_main)
            
        if i == 0: 
            df_total = df_chunks
        else:
            for ii, df_chunk in enumerate(df_chunks):
                df_total[ii] = pd.concat([df_total[ii], df_chunk])
        # k += 1
        if df_chunks[0] is not None:
            stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
                 df_chunks[0].shape[0], df_chunks[1].shape[0], df_chunks[2].shape[0]])
        else:
            stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
                 0, 0, 0])

    if df_total[0] is not None: 
        # print(df_total[0].shape)
        total_sheet_names = ['Услуги', 'ЛП', 'РМ' ]
        # fn_save = save_to_excel(df_total, total_sheet_names, path_tkbd_processed, 'tkbd.xlsx')
        fn_save = save_to_excel(df_total, total_sheet_names, data_processed_dir, 'tkbd_check.xlsx')
        # str_date = fn_save.replace('.xlsx', '').split('_')[-4:])
        # df_stat_tk = pd.DataFrame(stat_tk, columns = ['tk_profile', 'tk_code', 'tk_name', 'fn', 'sheet_name', 'Услуги', 'ЛП', 'РМ'])
        df_stat_tk = pd.DataFrame(stat_tk, columns = head_cols + ['Услуги', 'ЛП', 'РМ'])
        fm_stat_save = save_to_excel([df_stat_tk], 
                      ['Shapes'], data_processed_dir, 'tkbd_check_stat.xlsx')
    else: 
        fn_save = None
        fm_stat_save = None
    # logger.info(f"Check file '{fn_save}' saved in '{data_processed_dir}'")
    # logger.info(f"Check stat file '{fm_stat_save}' saved in '{data_processed_dir}'")
    logger.info(f"Файл проверки '{fn_save}' сохранен в '{data_processed_dir}'")
    logger.info(f"Файл статистики обработки '{fm_stat_save}' сохранен в '{data_processed_dir}'")

    return fn_save, fm_stat_save


# from pandas.io.excel._pyxlsb import PyxlsbReader
def run_check_by_files(data_source_dir, data_processed_dir,
                     print_debug = False, print_debug_main = True):
    df_total = [None, None, None]
    stat_tk = []
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel', 'Название листа в файле Excel']
    
    fn_lst = os.listdir(data_source_dir)
    k = 0
    fn_processed = []
    # for i, fn in tqdm(enumerate(fn_lst[:]), total = len(fn_lst)):
    for i, fn in tqdm(enumerate(fn_lst[10:]), total = len(fn_lst)):
    
        if not os.path.isfile(os.path.join(data_source_dir, fn)) or '.xls' not in fn.lower(): 
            # logger.info(f"file '{fn}' not found or not xlsx-file")
            logger.info(f"Файл '{fn}' не найден или не xls-файл")
            continue
        tk_profile = None
        tk_code = None
        tk_name = None #re.sub(r"^\d+\.", '', fn.split(' +')[0].replace('.xlsx','')).strip()
        patient_model = None
        
        # ws.sheet_state = 'hidden'
        fn_ext = fn.split('.')[-1].lower()
        if fn_ext in ['xlsx', 'xlsm', 'xls']:
            wb = load_workbook(os.path.join(data_source_dir, fn))
            xl_sheet_names = wb.get_sheet_names()
            xl_sheet_names = [ws_name for ws_name in xl_sheet_names if wb[ws_name].sheet_state != 'hidden']
        elif fn_ext in ['xlsb']:
            xl = pd.ExcelFile(os.path.join(data_source_dir, fn))
            # print("xl.engine:", xl.engine)
            
            # xl.engine == 'openpyxl'
            # sheet.title, sheet.sheet_state
            # print(dir(xl.book))
            # print(xl.book.sheets)
            # # print(xl.book._sheets)
            # sheets = xl.book.sheets
            # w_s = [xl.book.get_sheet(sheet) for sheet in sheets]
            # print(w_s)
            # print(dir(w_s[0]))
            # xl_sheet_names = [sheet for sheet in sheets if xl.book.get_sheet(sheet).sheet_state != 'hidden']
            
            # xl.engine == 'xlrd'
            # sheets = xl.book.sheets
            # xl_sheet_names = [sheet.name for sheet in sheets if sheet.visibility]
            
            # xl.engine: pyxlsb - нет атрибута hidden
            
            xl_sheet_names = xl.sheet_names  # see all sheet names
            # print(dir(xl))
        else: xl_sheet_names = []
        # # print(fn, xl_sheet_names)
        logger.info(f"Файл: '{fn}'")
        logger.info(f"Листы: {str(xl_sheet_names)}")
        fn_processed.append([fn, False])
        
        for sheet_name in xl_sheet_names:

            # if fn_ext == 'xlsb':
            #     try:
            #         df_tk = xl.parse(sheet_name) #                _reader = pd.io.excel._pyxlsb.PyxlsbReader
            #     except Exception as err:
            #         print(fn, sheet_name, err)
            # else:
            #     df_tk = pd.read_excel(os.path.join(data_source_dir, fn), sheet_name= sheet_name)
            df_tk = pd.read_excel(os.path.join(data_source_dir, fn), sheet_name= sheet_name)
            # ExcelFile.parse
            # print(k, sheet_name)
            logger.info(f"{k}: Лист: '{sheet_name}'")
    
            # logger.error('В описнаии нет названий листов Excel')
            # sys.exit(2)
        
        
            if print_debug_main: 
                print()
                print(fn, sheet_name)
            df_chunks = [None, None, None]
            try:
                df_chunks = run_check_TK_02(data_source_dir, data_processed_dir, fn, sheet_name,
                     tk_code, tk_profile, tk_name, patient_model,
                     exit_at_not_all_cols=False,
                     print_debug = print_debug, print_debug_main = print_debug_main)
            except Exception as err:
                logger.error(f"Файл '{fn}'")
                logger.error(f"Лист '{sheet_name}'")
                logger.error(f"Не обработан из-за ошибки: '{err}'")
                
            # if df_chunks[0] is None : continue
            
            if k == 0: 
                df_total = df_chunks
            else:
                for ii, df_chunk in enumerate(df_chunks):
                    if df_chunk is not None:
                        df_total[ii] = pd.concat([df_total[ii], df_chunk])
                    # else:
                    #     df_total[ii] = pd.concat([df_total[ii], pd.DataFrame([None], columns=[None])])
            k += 1
            
            
            row_stat = []
            for df_chunk in df_chunks:
                if df_chunk is not None:
                    row_stat.append(df_chunk.shape[0])
                    fn_processed[-1][1] = True # fn_processed.append([fn, False])
                else:
                    row_stat.append(0)
            stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
                         row_stat[0], row_stat[1], row_stat[2]])
            # if df_chunks[0] is not None:
            #     stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
            #          df_chunks[0].shape[0], df_chunks[1].shape[0], df_chunks[2].shape[0]])
            #     fn_processed[-1][1] = True # fn_processed.append([fn, False])
            # else:
            #     stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
            #          0, 0, 0])

    if df_total[0] is not None or df_total[1] is not None or df_total[2] is not None: 
        # print(df_total[0].shape)
        total_sheet_names = ['Услуги', 'ЛП', 'РМ' ]
        # df_total_to_write = []
        # total_sheet_names_to_write = []
        # for i_chunk, df_total_chunk in enumerate(df_total):
        #     if df_total_chunk is not None:
        #         df_total_to_write.append(df_total_chunk)
        #         # total_sheet_names_to_write.append(total_sheet_names[i_chunk])
        #     else: 
        #         df_total_to_write.append(pd.DataFrame([None], columns=[None]))
            # fn_save = save_to_excel(df_total, total_sheet_names, path_tkbd_processed, 'tkbd.xlsx')
        # fn_save = save_to_excel(df_total, total_sheet_names, data_processed_dir, 'tkbd_check.xlsx')
        fn_save = save_to_excel(df_total, total_sheet_names, data_processed_dir, 'tkbd_check.xlsx')

        # str_date = fn_save.replace('.xlsx', '').split('_')[-4:])
        # df_stat_tk = pd.DataFrame(stat_tk, columns = ['tk_profile', 'tk_code', 'tk_name', 'fn', 'sheet_name', 'Услуги', 'ЛП', 'РМ'])
        stat_tk_clean = [lst for lst in stat_tk if lst[-3:] != [0,0,0]]
        # print("stat_tk:", stat_tk[:2], stat_tk_clean[:2])
        df_stat_tk_clean = pd.DataFrame(stat_tk_clean, columns = head_cols + ['Услуги', 'ЛП', 'РМ'])
        df_stat_tk = pd.DataFrame(stat_tk, columns = head_cols + ['Услуги', 'ЛП', 'РМ'])
        df_stat_tk_files = pd.DataFrame(fn_processed, columns = ['Файл', 'Обработан'])

        fm_stat_save = save_to_excel([df_stat_tk_clean, df_stat_tk_files, df_stat_tk], 
                      ['Files_Sheets', 'Files', 'Files_Sheets_all'], data_processed_dir, 'tkbd_check_stat.xlsx')
    else: 
        fn_save = None
        fm_stat_save = None
    # logger.info(f"Check file '{fn_save}' saved in '{data_processed_dir}'")
    # logger.info(f"Check stat file '{fm_stat_save}' saved in '{data_processed_dir}'")
    logger.info(f"Файл проверки '{fn_save}' сохранен в '{data_processed_dir}'")
    logger.info(f"Файл статистики обработки '{fm_stat_save}' сохранен в '{data_processed_dir}'")
    
    return fn_save, fm_stat_save

def run_check_by_files_01(data_source_dir, data_processed_dir,
                     print_debug = False, print_debug_main = True):
    df_total = [None, None, None]
    stat_tk = []
    head_cols = ['Профиль', 'Код ТК', 'Наименование ТК', 'Модель пациента', 'Файл Excel', 'Название листа в файле Excel']
    
    fn_lst = os.listdir(data_source_dir)
    k = 0
    
    for i, fn in tqdm(enumerate(fn_lst[:]), total = len(fn_lst)):
    
        if not os.path.isfile(os.path.join(data_source_dir, fn)) or '.xls' not in fn.lower(): 
            # logger.info(f"file '{fn}' not found or not xlsx-file")
            logger.info(f"Файл '{fn}' не найден или не xls-файл")
            continue
        tk_profile = None
        tk_code = None
        tk_name = None #re.sub(r"^\d+\.", '', fn.split(' +')[0].replace('.xlsx','')).strip()
        patient_model = None
        xl = pd.ExcelFile(os.path.join(data_source_dir, fn))
        xl_sheet_names = xl.sheet_names  # see all sheet names
        # print(fn, xl_sheet_names)
        logger.info(f"Файл: '{fn}'")
        logger.info(f"Лист: {str(xl_sheet_names)}")
        for sheet_name in xl_sheet_names:

            df_tk = pd.read_excel(os.path.join(data_source_dir, fn), sheet_name= sheet_name)

            # print(k, sheet_name)
            logger.info(f"{k}: {sheet_name}")
    
            # logger.error('В описнаии нет названий листов Excel')
            # sys.exit(2)
        
        
            if print_debug_main: 
                print()
                print(fn, sheet_name)
            df_chunks = run_check_TK(data_source_dir, data_processed_dir, fn, sheet_name,
                 tk_code, tk_profile, tk_name, patient_model,
                 exit_at_not_all_cols=False,
                 print_debug = print_debug, print_debug_main = print_debug_main)
            if df_chunks[0] is None : continue
            
            if k == 0: 
                df_total = df_chunks
            else:
                for ii, df_chunk in enumerate(df_chunks):
                    df_total[ii] = pd.concat([df_total[ii], df_chunk])
            k += 1
            if df_chunks[0] is not None:
                stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
                     df_chunks[0].shape[0], df_chunks[1].shape[0], df_chunks[2].shape[0]])
            else:
                stat_tk.append( [tk_profile, tk_code, tk_name, patient_model, fn, sheet_name, 
                     0, 0, 0])

    if df_total[0] is not None: 
        print(df_total[0].shape)
        total_sheet_names = ['Услуги', 'ЛП', 'РМ' ]
        # fn_save = save_to_excel(df_total, total_sheet_names, path_tkbd_processed, 'tkbd.xlsx')
        fn_save = save_to_excel(df_total, total_sheet_names, data_processed_dir, 'tkbd_check.xlsx')
        # str_date = fn_save.replace('.xlsx', '').split('_')[-4:])
        # df_stat_tk = pd.DataFrame(stat_tk, columns = ['tk_profile', 'tk_code', 'tk_name', 'fn', 'sheet_name', 'Услуги', 'ЛП', 'РМ'])
        df_stat_tk = pd.DataFrame(stat_tk, columns = head_cols + ['Услуги', 'ЛП', 'РМ'])
        

        fm_stat_save = save_to_excel([df_stat_tk], 
                        ['Shapes'], data_processed_dir, 'tkbd_check_stat.xlsx')
    else: 
        fn_save = None
        fm_stat_save = None
    # logger.info(f"Check file '{fn_save}' saved in '{data_processed_dir}'")
    # logger.info(f"Check stat file '{fm_stat_save}' saved in '{data_processed_dir}'")
    logger.info(f"Файл проверки '{fn_save}' сохранен в '{data_processed_dir}'")
    logger.info(f"Файл статистики обработки '{fm_stat_save}' сохранен в '{data_processed_dir}'")

    return fn_save, fm_stat_save


def add_check_comments(path_tkbd_processed, fn_save):
    wb = load_workbook(os.path.join(path_tkbd_processed, fn_save))
    cols_wdth_lst = [[5,20,70,10,15,15,10,10], [5,20,15,25,15,15,10,10], [5,70,15,15,15,15,10,10]]
    desc_cols_num = 6 # Кол-во колонок перед: Профиль Код Наименование ТК Модель пациента Файл Лист
    col_num_check_row_total_lst = [8+desc_cols_num, 8+desc_cols_num, 7+desc_cols_num]
    # 8/7 кол-во колонок в каждом чанке перед результирующим кодом проверки 
    col_num_check_row_codes_lst = [9+desc_cols_num, 9+desc_cols_num, 8+desc_cols_num]
    
    # print(wb.sheetnames)
    for chunk_num, ws_title in enumerate(wb.sheetnames):
        ws = wb[ws_title] #wb['Услуги']
        # ws.delete_cols(col_num_check_row_total_lst[chunk_num], 40)
        
        # chunk_num = 0
        # if chunk_num==2: continue

        col_num_check_row_total = col_num_check_row_total_lst[chunk_num] #8
        col_num_check_row_codes = col_num_check_row_codes_lst[chunk_num] #9

        alignment=Alignment(horizontal='left', #'general',
                             vertical= 'top', #'bottom',
                             text_rotation=0,
                             wrap_text=True,
                             shrink_to_fit=False,
                             indent=0)
        cols_wdth = cols_wdth_lst[chunk_num] #[5,20,70,10,15,15,10,10]
        ws.auto_filter.ref = "A1:X1"
        for ir, row in enumerate(ws.values):
            # print(ir) #, row)
            # print(type(row[col_num_check_row_total]), row[col_num_check_row_total])
            # if (row[col_num_check_row_total]=='FALSE') and (ir>0):
            if ir==0:
                for ic, _ in enumerate(row):
                    cell = ws.cell(row=ir+1, column=ic+1 + desc_cols_num)
                    cell.comment = None
                    cell.alignment = alignment
                    if ic < len(cols_wdth):
                        ws.column_dimensions[cell.column_letter].width = cols_wdth[ic]
                    else: ws.column_dimensions[cell.column_letter].width = 15

            if (not row[col_num_check_row_total]) and (ir>0):
                # for ic,value in enumerate(row):
                #     print(ic, value)
                # print(type(row[col_num_check_row_codes]), row[col_num_check_row_codes])
                s = row[col_num_check_row_codes]
                # print("s:", s)
                # s = s.replace('[','').replace(']','').split()
                # print("s:", s)
                # rez_code_values = [int (ss) for ss in s]
                if s is not None:
                    str_lst = transform_list_form_xlsx(s)
                    # print(str_lst)
                    rez_code_values = conv_str_lst_2_int_lst(transform_list_form_xlsx(row[col_num_check_row_codes]))
                    # print(rez_code_values)
                    err_messages = get_err_messages(rez_code_values, err_msg_lst[chunk_num])
                    # print(err_messages)
                    for ic, err_msg_sl in enumerate(err_messages):
                        # print('->len', len(err_msg_sl))
                        comment = None
                        # cell = None
                        cell = ws.cell(row=ir+1, column=ic+1 + desc_cols_num)
                        if len(err_msg_sl)>0:
                            comment = Comment('\n'.join(err_msg_sl), "test")
                            # print(f"ic: {ic}, ir: {ir}")
                            # print(comment)
                            comment.width = 300
                            # comment.height = 50* len(err_msg_sl)
                            comment.height = 100
                            # ws["A1"].comment = comment
                            # cell = ws.cell(row=ir+1, column=ic+1)
                            # print(f"cell.coordinate: {cell.coordinate}")
                            cell.comment = comment
                            cell.fill = PatternFill('solid', fgColor="faf080")
                            # ws.cell(row=ir+1, column=ic+1, comment= comment)
                        else:
                            cell.comment = None
                            cell.fill = PatternFill('solid', fgColor="ffffff")
            # else:
            # if ir>20: break

        ws.delete_cols(col_num_check_row_codes_lst[chunk_num]+1, 40)

    offset = datetime.timezone(datetime.timedelta(hours=3))
    dt = datetime.datetime.now(offset)
    str_date = dt.strftime("%Y_%m_%d_%H%M")
    fn_ch_com_save = 'tkbd_check_commented_' + str_date + '.xlsx'
    wb.save(os.path.join(path_tkbd_processed, fn_ch_com_save))    
    # logger.info(f" file '{fn_ch_com_save}' save in '{path_tkbd_processed}'")
    # logger.info(f" file '{fn_ch_com_save}' save in '{path_tkbd_processed}'")
    logger.info(f"Файл с примечаниями '{fn_ch_com_save}' сохранен в '{path_tkbd_processed}'")





def check_input_pars(data_source_dir, data_processed_dir, data_root_dir, xlsx_description, supp_dict_dir) :
    if not os.path.exists(data_source_dir):
        logger.error(f"Not found directory for input TK xlsx-files: '{data_source_dir}'")
        sys.exit(2)
    if not os.path.exists(data_processed_dir):
        logger.error(f"Not found directory for processed TK xlsx-files: '{data_processed_dir}'")
        sys.exit(2)
    if data_root_dir is not None and not os.path.exists(data_root_dir):
        logger.error(f"Not found data root directory '{data_root_dir}'")
        sys.exit(2)
    if xlsx_description is not None \
        and not os.path.exists(os.path.join( data_root_dir,xlsx_description)):
        logger.error(f"Not found xlsx descrition TKЖ file '{xlsx_description}' in directory '{data_root_dir}'")
        sys.exit(2)
    if supp_dict_dir is not None and not os.path.exists(supp_dict_dir):
        logger.error(f"Not found directory for support dictionaries '{supp_dict_dir}'")
        sys.exit(2)
        
    return True

def parse_opt():
    parser = argparse.ArgumentParser()
    parser.add_argument('--data_source_dir', '-di', type=str, default='./data/source/',
        help="Directory for input TK xlsx-files, default  './data/source/'")
    parser.add_argument('--data_processed_dir', '-do', type=str, default='./data/processed/',
        help="Directory for processed TK xlsx-files, default  './data/processed/'")
    parser.add_argument('--data_root_dir', '-dr', type=str, default= None, # './data/',
        help="Dir for xlsx descrition TK file  (file< TK, sheet_name in data_root directory', default  'tk_descript.xlsx")
    parser.add_argument('--xlsx_description', '-desc', type=str, default= None, #'tk_descript.xlsx',
        help="Xlsx descrition TK file  (file_nmae, TK_name, TK_code, sheet_name)' in data root dir, default None")
    parser.add_argument('--supp_dict_dir', '-dd', type=str, default='./data/supp_dict/',
        help="Directory for support dictionaries, default  './data/supp_dict/'")
            
    opt = parser.parse_args()
    return opt

def main (data_source_dir = './data/source/',
    data_processed_dir = './data/processed/',
    data_root_dir = None, #'./data/',
    xlsx_description = None, 
    supp_dict_dir = './data/supp_dict/',
    ):
    # global smnn_list_df, klp_list_dict_df, zvnlp_df, znvlp_date, znvlp_date_format, esklp_date_format #esklp_date
    global df_services_MGFOMS, df_services_804n, df_RM, df_MNN, df_mi_org_gos, df_mi_national
    
    check_input_pars(data_source_dir, data_processed_dir, data_root_dir, xlsx_description, supp_dict_dir)    

    # supp_dict_dir =  'D:/DPP/02_TKBD/data/supp_dict/source/'
    # load_check_dictionaries(os.path.join(data_root_dir,'supp_dict'))
    # df_services_MGFOMS, df_services_804n, df_RM, df_MNN = load_check_dictionaries(supp_dict_dir)
    df_services_MGFOMS, df_services_804n, df_RM, df_MNN, df_mi_org_gos, df_mi_national = load_check_dictionaries(supp_dict_dir)

    if xlsx_description is None:
        # run_check_by_files(data_source_dir, data_processed_dir)
        fn_save, fm_stat_save = run_check_by_files(data_source_dir, data_processed_dir,
                     print_debug = False, print_debug_main = False) #True)
    else: # run_check_by_desc(data_source_dir, data_processed_dir, data_root_dir, xlsx_description)
        fn_save, fm_stat_save = run_check_by_desc(data_root_dir, xlsx_description, data_source_dir, data_processed_dir,
                     print_debug = False, print_debug_main = False) #True)
    
    add_check_comments(data_processed_dir, fn_save)    

    # print('check_code_MGFOMS:', check_code_MGFOMS('1001'))

if __name__ == '__main__':
    if len(sys.argv) > 1: # есть аргументы в командной строке
        opt = parse_opt()
        main(**vars(opt))
    else:
        main()

# jupyter notebook
# py tk_test.py -di "D:/DPP/02_TKBD/data/tk/source/" -do "D:/DPP/02_TKBD/data/tk/processed/" -dd "D:/DPP/02_TKBD/data/supp_dict/source/"
# py tk_test.py -di "D:/DPP/02_TKBD/data/tk/source/tk_2023_02_16" -do "D:/DPP/02_TKBD/data/tk/processed/" -dd "D:/DPP/02_TKBD/data/supp_dict/source/"

# colab
# !python tk_test.py -di "/content/data/source/" -do "/content/data/processed/" -dr "/content/data/" -dd "/content/data/supp_dict"
